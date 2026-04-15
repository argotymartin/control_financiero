#!/usr/bin/env python3
"""
App Web - Control Financiero
Version web completa con todas las funciones.
"""
from flask import (Flask, render_template_string, send_from_directory,
                   request, redirect, url_for, flash, send_file, session, jsonify)
from functools import wraps
import hashlib
from datetime import datetime
import json
import os
import re
import smtplib
import zipfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import base64
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import face_recognition
import numpy as np

RUTA_APP = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_DATOS = os.path.join(RUTA_APP, 'pagos_datos.json')
ARCHIVO_EXCEL = os.path.join(RUTA_APP, 'control_financiero.xlsx')
ARCHIVO_DESTINATARIOS = os.path.join(RUTA_APP, 'destinatarios.json')

ARCHIVO_USUARIOS = os.path.join(RUTA_APP, 'usuarios.json')
CARPETA_ROSTROS = os.path.join(RUTA_APP, 'rostros')
os.makedirs(CARPETA_ROSTROS, exist_ok=True)

CORREO_REMITENTE = os.environ.get('CORREO_REMITENTE', 'argoty.martin@gmail.com')
CORREO_CLAVE_APP = os.environ.get('CORREO_CLAVE_APP', 'vbzs yrtk auhr quya')

MESES = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
}

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'pagos_liliana_2024')


def hash_clave(clave):
    return hashlib.sha256(clave.encode()).hexdigest()


def cargar_usuarios():
    if os.path.exists(ARCHIVO_USUARIOS):
        with open(ARCHIVO_USUARIOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    usuario_admin = {
        'admin': {
            'nombre': 'Administrador',
            'clave': hash_clave('admin123')
        }
    }
    guardar_usuarios(usuario_admin)
    return usuario_admin


def guardar_usuarios(usuarios):
    with open(ARCHIVO_USUARIOS, 'w', encoding='utf-8') as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=2)


def guardar_imagen_base64(imagen_base64, ruta):
    datos_img = base64.b64decode(imagen_base64.split(',')[1])
    with open(ruta, 'wb') as f:
        f.write(datos_img)


def obtener_encoding_facial(ruta):
    img = face_recognition.load_image_file(ruta)
    encodings = face_recognition.face_encodings(img)
    return encodings[0] if encodings else None


def registrar_rostro(usuario, imagen_base64):
    ruta_foto = os.path.join(CARPETA_ROSTROS, f'{usuario}.jpg')
    guardar_imagen_base64(imagen_base64, ruta_foto)
    encoding = obtener_encoding_facial(ruta_foto)
    if encoding is None:
        os.remove(ruta_foto)
        return False
    usuarios = cargar_usuarios()
    if usuario in usuarios:
        usuarios[usuario]['rostro'] = encoding.tolist()
        guardar_usuarios(usuarios)
    return True


def verificar_rostro(imagen_base64):
    ruta_temp = os.path.join(CARPETA_ROSTROS, '_temp_login.jpg')
    guardar_imagen_base64(imagen_base64, ruta_temp)
    encoding_login = obtener_encoding_facial(ruta_temp)
    os.remove(ruta_temp)
    if encoding_login is None:
        return None
    usuarios = cargar_usuarios()
    for usuario, datos in usuarios.items():
        if 'rostro' in datos:
            encoding_guardado = np.array(datos['rostro'])
            resultado = face_recognition.compare_faces([encoding_guardado], encoding_login, tolerance=0.6)
            if resultado[0]:
                return usuario
    return None


def login_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        if 'usuario' not in session:
            flash('Debes iniciar sesion para acceder.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorador


def admin_requerido(f):
    @wraps(f)
    @login_requerido
    def decorador(*args, **kwargs):
        if session.get('usuario') != 'admin':
            flash('Solo el administrador puede acceder a esta funcion.', 'error')
            return redirect(url_for('inicio'))
        return f(*args, **kwargs)
    return decorador


def cargar_pagos():
    if os.path.exists(ARCHIVO_DATOS):
        with open(ARCHIVO_DATOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def guardar_pagos(pagos):
    with open(ARCHIVO_DATOS, 'w', encoding='utf-8') as f:
        json.dump(pagos, f, ensure_ascii=False, indent=2)


def cargar_destinatarios():
    if os.path.exists(ARCHIVO_DESTINATARIOS):
        with open(ARCHIVO_DESTINATARIOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    return ['martin.argoty@hotmail.com', 'lipamoan@hotmail.com']


def guardar_destinatarios(lista):
    with open(ARCHIVO_DESTINATARIOS, 'w', encoding='utf-8') as f:
        json.dump(lista, f, ensure_ascii=False, indent=2)


def extraer_datos_imagen(ruta_imagen):
    img = Image.open(ruta_imagen)
    texto = pytesseract.image_to_string(img, lang='spa')

    datos = {'fecha': '', 'valor': 0, 'medio': '', 'referencia': '', 'observacion': ''}

    es_nequi = 'nequi' in texto.lower() or 'envío realizado' in texto.lower() or 'envio recibido' in texto.lower() or 'envío recibido' in texto.lower()
    es_banco_bogota = 'banco de bogotá' in texto.lower() or 'banco de bogota' in texto.lower()
    es_transfiya = 'transfiya' in texto.lower()

    patron_valor = re.search(r'\$\s*([\d.,]+)', texto)
    if patron_valor:
        val_str = patron_valor.group(1).replace('.', '').replace(',', '')
        if len(val_str) > 2 and val_str.endswith('00'):
            val_str = val_str[:-2]
        try:
            datos['valor'] = int(val_str)
        except ValueError:
            pass

    patron_fecha_nequi = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})', texto, re.IGNORECASE)
    patron_fecha_banco = re.search(r'(\w+)\s+(\d{1,2})\s+del\s+(\d{4})', texto, re.IGNORECASE)
    patron_fecha_alt = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+del\s+(\d{4})', texto, re.IGNORECASE)

    if patron_fecha_nequi:
        dia = patron_fecha_nequi.group(1).zfill(2)
        mes = MESES.get(patron_fecha_nequi.group(2).lower(), '00')
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_nequi.group(3)}'
    elif patron_fecha_banco:
        mes = MESES.get(patron_fecha_banco.group(1).lower(), '00')
        dia = patron_fecha_banco.group(2).zfill(2)
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_banco.group(3)}'
    elif patron_fecha_alt:
        dia = patron_fecha_alt.group(1).zfill(2)
        mes = MESES.get(patron_fecha_alt.group(2).lower(), '00')
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_alt.group(3)}'

    if es_nequi and es_transfiya:
        datos['medio'] = 'Banco de Bogota / Transfiya'
    elif es_nequi and es_banco_bogota:
        datos['medio'] = 'Banco de Bogota / Nequi'
    elif es_nequi:
        datos['medio'] = 'Nequi'
    elif es_transfiya:
        datos['medio'] = 'Banco de Bogota / Transfiya'
    elif es_banco_bogota:
        datos['medio'] = 'Banco de Bogota'
    else:
        datos['medio'] = 'Otro'

    patron_ref_nequi = re.search(r'[Rr]eferencia\s*\n?\s*(\S+)', texto)
    patron_ref_banco = re.search(r'autorizaci[oó]n[:\s]*\n?\s*(\S+)', texto, re.IGNORECASE)
    patron_comprobante = re.search(r'[Cc]omprobante\s*\n?\s*(\d{10,})', texto)

    if patron_ref_nequi:
        ref = re.sub(r'^[^A-Za-z0-9]+', '', patron_ref_nequi.group(1).strip())
        datos['referencia'] = ref
    elif patron_ref_banco:
        datos['referencia'] = patron_ref_banco.group(1).strip()
    elif patron_comprobante:
        datos['referencia'] = patron_comprobante.group(1).strip()

    patron_para = re.search(r'(?:Para|De)\s*\n\s*(.+)', texto)
    patron_destino = re.search(r'[Cc]uenta destino[:\s]*\n?\s*(.+)', texto)
    patron_nota = re.search(r'[Nn]ota[:\s]*\n?\s*(.+)', texto)

    obs_partes = []
    if patron_para:
        obs_partes.append(patron_para.group(1).strip())
    if patron_destino:
        obs_partes.append('Destino: ' + patron_destino.group(1).strip())
    if patron_nota:
        obs_partes.append('Nota: ' + patron_nota.group(1).strip())
    datos['observacion'] = ' - '.join(obs_partes) if obs_partes else ''

    return datos


def generar_excel(pagos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Financiero"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    link_font = Font(color="0563C1", underline="single")
    rojo = Font(color="DC3545")
    verde = Font(color="28A745")

    ws.merge_cells('A1:J1')
    ws['A1'] = 'Control Financiero'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    if pagos:
        fecha_desde = pagos[0].get('fecha', '')
        fecha_hasta = pagos[-1].get('fecha', '')
        ws['A2'] = f'Desde: {fecha_desde}  -  Hasta: {fecha_hasta}'
    else:
        ws['A2'] = ''
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No.', 'Fecha', 'Concepto', 'Debito', 'Credito', 'Saldo',
               'Medio de Pago', 'Referencia', 'Observaciones', 'Comprobante']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    saldo = 0
    total_debitos = 0
    total_creditos = 0
    for i, pago in enumerate(pagos):
        row = i + 5
        tipo = pago.get('tipo', 'egreso')
        valor = pago['valor']
        debito = valor if tipo == 'egreso' else 0
        credito = valor if tipo == 'ingreso' else 0
        saldo += credito - debito
        total_debitos += debito
        total_creditos += credito

        ws.cell(row=row, column=1, value=i + 1).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=pago['fecha']).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=pago.get('concepto', '')).border = border

        cell_deb = ws.cell(row=row, column=4, value=debito if debito else None)
        cell_deb.border = border
        cell_deb.number_format = '$#,##0'
        cell_deb.alignment = Alignment(horizontal='right')
        if debito:
            cell_deb.font = rojo

        cell_cred = ws.cell(row=row, column=5, value=credito if credito else None)
        cell_cred.border = border
        cell_cred.number_format = '$#,##0'
        cell_cred.alignment = Alignment(horizontal='right')
        if credito:
            cell_cred.font = verde

        cell_saldo = ws.cell(row=row, column=6, value=saldo)
        cell_saldo.border = border
        cell_saldo.number_format = '$#,##0'
        cell_saldo.alignment = Alignment(horizontal='right')
        cell_saldo.font = verde if saldo >= 0 else rojo

        ws.cell(row=row, column=7, value=pago['medio']).border = border
        ws.cell(row=row, column=8, value=pago['referencia']).border = border
        ws.cell(row=row, column=9, value=pago.get('observacion', '')).border = border

        if pago.get('imagen'):
            cell_link = ws.cell(row=row, column=10, value='Ver comprobante')
            cell_link.hyperlink = './' + pago['imagen']
            cell_link.font = link_font
            cell_link.alignment = Alignment(horizontal='center')
            cell_link.border = border

    total_row = 5 + len(pagos)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=11)
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).border = border

    cell_td = ws.cell(row=total_row, column=4, value=total_debitos)
    cell_td.number_format = '$#,##0'
    cell_td.font = Font(bold=True, size=11, color="DC3545")
    cell_td.alignment = Alignment(horizontal='right')

    cell_tc = ws.cell(row=total_row, column=5, value=total_creditos)
    cell_tc.number_format = '$#,##0'
    cell_tc.font = Font(bold=True, size=11, color="28A745")
    cell_tc.alignment = Alignment(horizontal='right')

    cell_ts = ws.cell(row=total_row, column=6, value=saldo)
    cell_ts.number_format = '$#,##0'
    cell_ts.font = Font(bold=True, size=11)
    cell_ts.alignment = Alignment(horizontal='right')

    ws.cell(row=total_row, column=9, value=f'{len(pagos)} movimientos')

    anchos = {'A': 8, 'B': 14, 'C': 22, 'D': 16, 'E': 16, 'F': 16,
              'G': 28, 'H': 28, 'I': 40, 'J': 18}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(ARCHIVO_EXCEL)
    return saldo


# ==================== PLANTILLAS HTML ====================

HTML_HEAD = ('<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
             '<meta name="viewport" content="width=device-width, initial-scale=1.0">')
HTML_FOOTER = '<div class="footer">Control Financiero</div></body></html>'

SCRIPT_TOGGLE_CLAVE = """
<script>
function toggleClave(id, btn) {
    var input = document.getElementById(id);
    if (input.type === 'password') {
        input.type = 'text';
        btn.innerHTML = '&#128064;';
    } else {
        input.type = 'password';
        btn.innerHTML = '&#128065;';
    }
}
</script>
"""

ESTILOS = """
<style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { font-family: 'Segoe UI', Arial, sans-serif; background: #F0F2F5; color: #333; }
    .header {
        background: linear-gradient(135deg, #2E75B6, #1A4A7A);
        color: white; padding: 25px 30px; text-align: center;
        box-shadow: 0 2px 10px rgba(0,0,0,0.2);
    }
    .cabecera-fija {
        position: sticky; top: 0; z-index: 100;
    }
    .header h1 { font-size: 24px; margin-bottom: 5px; }
    .header p { font-size: 14px; opacity: 0.85; }
    .mensaje {
        max-width: 800px; margin: 15px auto; padding: 12px 20px;
        border-radius: 8px; font-size: 14px; text-align: center;
    }
    .mensaje-exito { background: #D4EDDA; color: #155724; border: 1px solid #C3E6CB; }
    .mensaje-error { background: #F8D7DA; color: #721C24; border: 1px solid #F5C6CB; }
    .mensaje-info { background: #D1ECF1; color: #0C5460; border: 1px solid #BEE5EB; }
    .barra-acciones {
        display: flex; justify-content: center; gap: 12px;
        padding: 15px 20px; flex-wrap: wrap;
        background: #F0F2F5;
    }
    .btn {
        display: inline-block; padding: 10px 20px; border: none; border-radius: 6px;
        font-size: 14px; font-weight: bold; cursor: pointer; text-decoration: none;
        color: white; transition: opacity 0.2s;
    }
    .btn:hover { opacity: 0.85; }
    .btn-azul { background: #0078D4; }
    .btn-verde { background: #217346; }
    .btn-rojo { background: #D44638; }
    .resumen { display: flex; justify-content: center; gap: 30px; padding: 20px; flex-wrap: wrap; }
    .tarjeta {
        background: white; border-radius: 10px; padding: 20px 30px;
        text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); min-width: 180px;
    }
    .tarjeta .numero { font-size: 28px; font-weight: bold; color: #2E75B6; }
    .tarjeta .etiqueta { font-size: 13px; color: #777; margin-top: 5px; }
    .contenedor-tabla {
        padding: 0 20px 30px; overflow: auto;
        max-height: 60vh;
    }
    table {
        width: 100%; border-collapse: collapse; background: white;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    thead th {
        background: #2E75B6; color: white; padding: 14px 12px;
        font-size: 13px; text-align: center; white-space: nowrap;
        border: 1px solid #2468A0;
        position: sticky; top: 0; z-index: 10;
    }
    tbody td {
        padding: 12px; text-align: center; vertical-align: middle;
        border: 1px solid #D0D0D0; font-size: 14px;
    }
    tbody tr:hover { background: #E8F4FD; }
    tbody tr:nth-child(even) { background: #F8F9FA; }
    tbody tr:nth-child(even):hover { background: #E8F4FD; }
    .valor { text-align: right; font-weight: 500; }
    .thumb-container {
        width: 80px; height: 60px; margin: 0 auto; overflow: hidden;
        border-radius: 5px; cursor: pointer; border: 2px solid #DDD; transition: border-color 0.2s;
    }
    .thumb-container:hover { border-color: #2E75B6; }
    .thumb-container img { width: 100%; height: 100%; object-fit: cover; }
    .sin-imagen {
        width: 80px; height: 60px; margin: 0 auto; background: #F0F0F0;
        border-radius: 5px; display: flex; align-items: center; justify-content: center;
        color: #AAA; font-size: 12px;
    }
    tfoot { position: sticky; bottom: 0; z-index: 10; }
    .fila-total { background: #2E75B6 !important; color: white; font-weight: bold; }
    .fila-total td {
        color: white; background: #2E75B6;
        padding: 14px 12px; border: 1px solid #2468A0;
    }
    .btn-eliminar-fila {
        background: #CC0000; color: white; border: none;
        padding: 5px 10px; border-radius: 4px; cursor: pointer; font-size: 12px;
    }
    .btn-eliminar-fila:hover { background: #990000; }
    .modal {
        display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%;
        background: rgba(0,0,0,0.85); z-index: 1000;
        justify-content: center; align-items: center; cursor: pointer;
    }
    .modal.activo { display: flex; }
    .modal img { max-width: 90%; max-height: 90%; border-radius: 8px; box-shadow: 0 0 30px rgba(0,0,0,0.5); }
    .modal .cerrar {
        position: absolute; top: 20px; right: 30px;
        color: white; font-size: 36px; cursor: pointer; font-weight: bold;
    }
    .modal .cerrar:hover { color: #FF6B6B; }
    .panel {
        max-width: 900px; margin: 0 auto; background: white;
        border-radius: 10px; padding: 25px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .panel h2 { margin-bottom: 15px; color: #2E75B6; font-size: 18px; }
    .form-grupo { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; margin-bottom: 15px; }
    .form-campo label { display: block; font-size: 13px; font-weight: bold; color: #555; margin-bottom: 4px; }
    .form-campo input {
        width: 100%; padding: 8px 12px; border: 1px solid #CCC;
        border-radius: 5px; font-size: 14px;
    }
    .form-campo input:focus { outline: none; border-color: #2E75B6; box-shadow: 0 0 4px rgba(46,117,182,0.3); }
    .lista-correos {
        list-style: none; padding: 0; max-height: 200px; overflow-y: auto;
        border: 1px solid #DDD; border-radius: 5px; margin: 10px 0;
    }
    .lista-correos li {
        display: flex; justify-content: space-between; align-items: center;
        padding: 8px 12px; border-bottom: 1px solid #EEE;
    }
    .lista-correos li:last-child { border-bottom: none; }
    .fila-agregar { display: flex; gap: 8px; margin-top: 10px; }
    .fila-agregar input { flex: 1; padding: 8px 12px; border: 1px solid #CCC; border-radius: 5px; }
    .footer {
        text-align: center; padding: 15px; color: #999; font-size: 12px;
        background: #F0F2F5; position: sticky; bottom: 0; z-index: 100;
        border-top: 1px solid #DDD;
    }
    .avatar {
        width: 36px; height: 36px; border-radius: 50%; object-fit: cover;
        border: 2px solid #AAD4FF; vertical-align: middle;
    }
    .avatar-iniciales {
        width: 36px; height: 36px; border-radius: 50%; display: inline-flex;
        align-items: center; justify-content: center; background: #AAD4FF;
        color: #1A4A7A; font-weight: bold; font-size: 14px; vertical-align: middle;
    }
    .barra-usuario {
        display: flex; align-items: center; justify-content: flex-end;
        gap: 10px; padding: 8px 20px; background: #1A4A7A;
    }
    @media (max-width: 768px) {
        .header h1 { font-size: 18px; }
        .resumen { gap: 10px; }
        .tarjeta { padding: 12px 18px; min-width: 140px; }
        .tarjeta .numero { font-size: 20px; }
        thead th, tbody td { padding: 8px 6px; font-size: 12px; }
        .thumb-container { width: 60px; height: 45px; }
        .sin-imagen { width: 60px; height: 45px; }
        .form-grupo { grid-template-columns: 1fr; }
        .barra-acciones { gap: 8px; }
        .btn { padding: 8px 14px; font-size: 12px; }
    }
</style>
"""

CABECERA = """
<div class="cabecera-fija">
<div class="header">
    <h1>Control Financiero</h1>
    {% if fecha_desde %}
    <p>Desde: {{ fecha_desde }} - Hasta: {{ fecha_hasta }}</p>
    {% endif %}
</div>
{% with mensajes = get_flashed_messages(with_categories=true) %}
{% if mensajes %}
{% for categoria, mensaje in mensajes %}
<div class="mensaje mensaje-{{ categoria }}">{{ mensaje }}</div>
{% endfor %}
{% endif %}
{% endwith %}
<div class="barra-usuario">
    <img src="/foto-usuario/{{ session.get('usuario', '') }}" class="avatar" style="cursor:pointer;"
         onclick="document.getElementById('modal-avatar').style.display='flex';"
         onerror="this.style.display='none'; this.nextElementSibling.style.display='inline-flex';">
    <span class="avatar-iniciales" style="display:none;">{{ session.get('nombre', ' ')[0] | upper }}</span>
    <span style="color: #AAD4FF; font-size: 13px;">
        {{ session.get('nombre', '') }}
    </span>
    <a href="/logout" style="color: #FF9999; font-size: 13px; text-decoration: none;">
        Cerrar Sesion
    </a>
</div>
<div class="barra-acciones">
    <a href="/" class="btn btn-azul">Inicio</a>
    <a href="/nuevo" class="btn btn-azul">Agregar Movimiento</a>
    <a href="/descargar-excel" class="btn btn-verde">Descargar Excel</a>
    <a href="/correo" class="btn btn-rojo">Enviar por Correo</a>
    {% if session.get('usuario') == 'admin' %}
    <a href="/usuarios" class="btn" style="background: #6C3483;">Usuarios</a>
    {% endif %}
</div>
</div>
<div id="modal-avatar" onclick="this.style.display='none';"
     style="display:none; position:fixed; top:0; left:0; width:100%; height:100%;
     background:rgba(0,0,0,0.85); z-index:1000; justify-content:center; align-items:center; cursor:pointer;">
    <img src="/foto-usuario/{{ session.get('usuario', '') }}"
         style="max-width:350px; max-height:350px; border-radius:50%; border:4px solid white;
         box-shadow:0 0 30px rgba(0,0,0,0.5); object-fit:cover;">
</div>
"""

MODAL_SCRIPT = """
<div class="modal" id="modal" onclick="cerrarModal()">
    <span class="cerrar">&times;</span>
    <img id="modal-img" src="" alt="Comprobante">
</div>
<script>
    function verImagen(src) {
        document.getElementById('modal-img').src = src;
        document.getElementById('modal').classList.add('activo');
    }
    function cerrarModal() {
        document.getElementById('modal').classList.remove('activo');
    }
    document.addEventListener('keydown', function(e) {
        if (e.key === 'Escape') cerrarModal();
    });
</script>
"""

PLANTILLA_LOGIN = (
    HTML_HEAD + '<title>Iniciar Sesion - Control Financiero</title>' + ESTILOS +
    """
    <style>
        .login-contenedor {
            display: flex; justify-content: center; align-items: center;
            min-height: 80vh; padding: 20px;
        }
        .login-panel {
            background: white; border-radius: 12px; padding: 40px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.15); width: 100%; max-width: 400px;
            text-align: center;
        }
        .login-panel h2 { color: #2E75B6; margin-bottom: 8px; font-size: 22px; }
        .login-panel p { color: #777; font-size: 14px; margin-bottom: 25px; }
        .login-campo { margin-bottom: 18px; text-align: left; }
        .login-campo label { display: block; font-size: 13px; font-weight: bold; color: #555; margin-bottom: 5px; }
        .login-campo input {
            width: 100%; padding: 12px 15px; border: 1px solid #CCC;
            border-radius: 8px; font-size: 15px;
        }
        .login-campo input:focus { outline: none; border-color: #2E75B6; box-shadow: 0 0 6px rgba(46,117,182,0.3); }
        .login-btn {
            width: 100%; padding: 14px; background: linear-gradient(135deg, #2E75B6, #1A4A7A);
            color: white; border: none; border-radius: 8px; font-size: 16px;
            font-weight: bold; cursor: pointer; transition: opacity 0.2s; margin-top: 10px;
        }
        .login-btn:hover { opacity: 0.9; }
        .login-icono { font-size: 50px; margin-bottom: 10px; }
    </style>
    </head><body>
    <div class="header">
        <h1>Control Financiero</h1>
    </div>
    {% with mensajes = get_flashed_messages(with_categories=true) %}
    {% if mensajes %}
    {% for categoria, mensaje in mensajes %}
    <div class="mensaje mensaje-{{ categoria }}">{{ mensaje }}</div>
    {% endfor %}
    {% endif %}
    {% endwith %}
    <div class="login-contenedor">
        <div class="login-panel">
            <div class="login-icono">&#128274;</div>
            <h2>Iniciar Sesion</h2>
            <p>Ingresa tus credenciales o usa tu rostro</p>

            <!-- Login clasico -->
            <div id="login-clasico">
                <form action="/login" method="post">
                    <div class="login-campo">
                        <label>Usuario:</label>
                        <input type="text" name="usuario" placeholder="Tu usuario" required autofocus>
                    </div>
                    <div class="login-campo">
                        <label>Contrasena:</label>
                        <div style="position: relative;">
                            <input type="password" name="clave" id="clave-login" placeholder="Tu contrasena" required style="padding-right: 45px;">
                            <span onclick="toggleClave('clave-login', this)" style="position: absolute; right: 12px; top: 50%; transform: translateY(-50%); cursor: pointer; font-size: 20px; user-select: none;">&#128065;</span>
                        </div>
                    </div>
                    <button type="submit" class="login-btn">INGRESAR</button>
                </form>
            </div>

            <!-- Separador -->
            <div style="display: flex; align-items: center; margin: 20px 0;">
                <hr style="flex: 1; border: none; border-top: 1px solid #DDD;">
                <span style="padding: 0 15px; color: #999; font-size: 13px;">o</span>
                <hr style="flex: 1; border: none; border-top: 1px solid #DDD;">
            </div>

            <!-- Login facial -->
            <div id="login-facial">
                <button type="button" class="login-btn" id="btn-login-facial" onclick="iniciarLoginFacial()"
                        style="background: linear-gradient(135deg, #28A745, #1E7E34);">
                    &#128247; INGRESAR CON MI CARA
                </button>
                <div id="zona-camara" style="display: none; margin-top: 15px;">
                    <video id="camara-login" width="280" height="210" style="border-radius: 8px; border: 3px solid #28A745; display: block; margin: 0 auto;"></video>
                    <canvas id="canvas-login" style="display: none;"></canvas>
                    <p id="estado-facial" style="font-size: 13px; color: #28A745; margin-top: 10px;">Mirate a la camara...</p>
                    <button type="button" class="btn btn-rojo" onclick="cancelarLoginFacial()" style="margin-top: 10px; font-size: 13px;">
                        Cancelar
                    </button>
                </div>
            </div>
        </div>
    </div>
    """ + SCRIPT_TOGGLE_CLAVE + """
    <script>
    var streamLogin = null;
    var intentoFacial = null;

    function iniciarLoginFacial() {
        var video = document.getElementById('camara-login');
        navigator.mediaDevices.getUserMedia({ video: true })
        .then(function(stream) {
            streamLogin = stream;
            video.srcObject = stream;
            video.play();
            document.getElementById('zona-camara').style.display = 'block';
            document.getElementById('btn-login-facial').style.display = 'none';
            document.getElementById('estado-facial').textContent = 'Mirate a la camara...';
            document.getElementById('estado-facial').style.color = '#28A745';
            setTimeout(capturarYVerificar, 2000);
        })
        .catch(function(err) {
            alert('No se pudo acceder a la camara: ' + err.message);
        });
    }

    function capturarYVerificar() {
        var video = document.getElementById('camara-login');
        var canvas = document.getElementById('canvas-login');
        canvas.width = video.videoWidth;
        canvas.height = video.videoHeight;
        var ctx = canvas.getContext('2d');
        ctx.drawImage(video, 0, 0);
        var dataUrl = canvas.toDataURL('image/jpeg', 0.8);

        document.getElementById('estado-facial').textContent = 'Verificando rostro...';
        document.getElementById('estado-facial').style.color = '#2E75B6';

        fetch('/login-facial', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ foto: dataUrl })
        })
        .then(function(r) { return r.json(); })
        .then(function(data) {
            if (data.ok) {
                document.getElementById('estado-facial').textContent = 'Bienvenido(a), ' + data.nombre + '!';
                document.getElementById('estado-facial').style.color = '#28A745';
                detenerCamara();
                setTimeout(function() { window.location.href = '/'; }, 1000);
            } else {
                document.getElementById('estado-facial').textContent = data.error + ' - Reintentando...';
                document.getElementById('estado-facial').style.color = '#DC3545';
                intentoFacial = setTimeout(capturarYVerificar, 2000);
            }
        })
        .catch(function() {
            document.getElementById('estado-facial').textContent = 'Error de conexion';
            document.getElementById('estado-facial').style.color = '#DC3545';
        });
    }

    function cancelarLoginFacial() {
        if (intentoFacial) clearTimeout(intentoFacial);
        detenerCamara();
        document.getElementById('zona-camara').style.display = 'none';
        document.getElementById('btn-login-facial').style.display = 'block';
    }

    function detenerCamara() {
        if (streamLogin) {
            streamLogin.getTracks().forEach(function(t) { t.stop(); });
            streamLogin = null;
        }
    }
    </script>
    """ + HTML_FOOTER
)

PLANTILLA_USUARIOS = (
    HTML_HEAD + '<title>Usuarios - Control Financiero</title>' + ESTILOS + '</head><body>'
    + CABECERA +
    """
    <div style="padding: 20px;">
        <div class="panel">
            <h2>Gestionar Usuarios</h2>

            <h3 style="font-size: 15px; margin-top: 15px; color: #555;">Crear nuevo usuario</h3>
            <form action="/crear-usuario" method="post" id="form-usuario">
                <div class="form-grupo">
                    <div class="form-campo">
                        <label>Usuario:</label>
                        <input type="text" name="usuario" placeholder="nombre de usuario" required>
                    </div>
                    <div class="form-campo">
                        <label>Nombre completo:</label>
                        <input type="text" name="nombre" placeholder="Nombre para mostrar" required>
                    </div>
                    <div class="form-campo">
                        <label>Contrasena:</label>
                        <div style="position: relative;">
                            <input type="password" name="clave" id="clave-nuevo" placeholder="Minimo 4 caracteres" required style="padding-right: 45px;">
                            <span onclick="toggleClave('clave-nuevo', this)" style="position: absolute; right: 12px; top: 50%; transform: translateY(-50%); cursor: pointer; font-size: 20px; user-select: none;">&#128065;</span>
                        </div>
                    </div>
                </div>
                <div style="margin-top: 15px; padding: 15px; border: 2px dashed #2E75B6; border-radius: 8px; text-align: center;">
                    <label style="font-size: 14px; font-weight: bold; color: #555;">Registro facial (opcional):</label>
                    <div style="margin-top: 10px;">
                        <video id="camara-registro" width="240" height="180" style="display:none; border-radius: 8px; border: 2px solid #2E75B6; margin: 0 auto;"></video>
                        <canvas id="canvas-registro" width="240" height="180" style="display:none;"></canvas>
                        <img id="foto-preview" style="display:none; border-radius: 8px; border: 2px solid #28A745; margin: 0 auto; width: 240px; height: 180px; object-fit: cover;">
                    </div>
                    <input type="hidden" name="foto_rostro" id="foto-rostro-input">
                    <div style="margin-top: 10px;">
                        <button type="button" class="btn btn-azul" onclick="activarCamara()" id="btn-activar-cam">Activar Camara</button>
                        <button type="button" class="btn btn-verde" onclick="capturarFoto()" id="btn-capturar" style="display:none;">Tomar Foto</button>
                        <button type="button" class="btn btn-rojo" onclick="reiniciarFoto()" id="btn-reiniciar" style="display:none;">Repetir</button>
                    </div>
                    <p id="estado-foto" style="font-size: 12px; color: #777; margin-top: 8px;">Sin foto registrada</p>
                </div>
                <button type="submit" class="btn btn-verde" style="margin-top: 15px;">Crear Usuario</button>
            </form>

            <h3 style="font-size: 15px; margin-top: 25px; color: #555;">Usuarios registrados</h3>
            <ul class="lista-correos">
                {% for usuario, datos in usuarios.items() %}
                <li>
                    <span>
                        <strong>{{ usuario }}</strong> - {{ datos.nombre }}
                        {% if usuario == 'admin' %}
                        <span style="color: #6C3483; font-size: 12px;">(administrador)</span>
                        {% endif %}
                        {% if datos.rostro is defined and datos.rostro %}
                        <span style="color: #28A745; font-size: 12px;">&#9989; rostro registrado</span>
                        {% else %}
                        <span style="color: #999; font-size: 12px;">&#10060; sin rostro</span>
                        {% endif %}
                    </span>
                    <span>
                        <button type="button" class="btn btn-azul" style="font-size: 11px; padding: 4px 10px;"
                                onclick="registrarRostroUsuario('{{ usuario }}')">
                            &#128247; {{ 'Actualizar rostro' if datos.rostro is defined and datos.rostro else 'Registrar rostro' }}
                        </button>
                        {% if usuario != 'admin' %}
                        <form action="/eliminar-usuario" method="post" style="display:inline;"
                              onsubmit="return confirm('Eliminar usuario {{ usuario }}?')">
                            <input type="hidden" name="usuario" value="{{ usuario }}">
                            <button type="submit" class="btn-eliminar-fila">Eliminar</button>
                        </form>
                        {% endif %}
                    </span>
                </li>
                {% endfor %}
            </ul>
        </div>
    </div>

    <!-- Modal para registrar rostro de usuario existente -->
    <div id="modal-rostro" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%;
         background:rgba(0,0,0,0.7); z-index:1000; justify-content:center; align-items:center;">
        <div style="background:white; border-radius:12px; padding:30px; text-align:center; max-width:400px; width:90%;">
            <h3 style="color:#2E75B6; margin-bottom:5px;">Registro Facial</h3>
            <p style="font-size: 18px; font-weight: bold; color: #333; margin-bottom: 15px;" id="modal-rostro-user"></p>
            <video id="camara-modal" width="280" height="210" style="border-radius:8px; border:3px solid #2E75B6; margin:0 auto;"></video>
            <canvas id="canvas-modal" style="display:none;"></canvas>
            <p id="estado-modal" style="font-size:13px; color:#2E75B6; margin:10px 0;">Ponte frente a la camara</p>
            <div style="margin-top:10px; display:flex; gap:10px; justify-content:center;">
                <button type="button" class="btn btn-verde" style="font-size:15px; padding:12px 30px;" onclick="capturarYGuardar()">
                    &#128247; Tomar Foto
                </button>
                <button type="button" class="btn btn-rojo" style="font-size:15px; padding:12px 30px;" onclick="cerrarModalRostro()">
                    Cancelar
                </button>
            </div>
        </div>
    </div>

    """ + SCRIPT_TOGGLE_CLAVE + """
    <script>
    var streamRegistro = null;
    var streamModal = null;
    var usuarioModal = '';

    function activarCamara() {
        var video = document.getElementById('camara-registro');
        navigator.mediaDevices.getUserMedia({ video: true })
        .then(function(stream) {
            streamRegistro = stream;
            video.srcObject = stream;
            video.play();
            video.style.display = 'block';
            document.getElementById('btn-activar-cam').style.display = 'none';
            document.getElementById('btn-capturar').style.display = 'inline-block';
            document.getElementById('estado-foto').textContent = 'Camara activa - ponte frente a la camara';
        })
        .catch(function(err) {
            alert('No se pudo acceder a la camara: ' + err.message);
        });
    }

    function capturarFoto() {
        var video = document.getElementById('camara-registro');
        var canvas = document.getElementById('canvas-registro');
        var ctx = canvas.getContext('2d');
        canvas.width = video.videoWidth;
        canvas.height = video.videoHeight;
        ctx.drawImage(video, 0, 0);
        var dataUrl = canvas.toDataURL('image/jpeg', 0.8);
        document.getElementById('foto-rostro-input').value = dataUrl;
        document.getElementById('foto-preview').src = dataUrl;
        document.getElementById('foto-preview').style.display = 'block';
        video.style.display = 'none';
        if (streamRegistro) {
            streamRegistro.getTracks().forEach(function(t) { t.stop(); });
        }
        document.getElementById('btn-capturar').style.display = 'none';
        document.getElementById('btn-reiniciar').style.display = 'inline-block';
        document.getElementById('estado-foto').textContent = 'Foto capturada correctamente';
        document.getElementById('estado-foto').style.color = '#28A745';
    }

    function reiniciarFoto() {
        document.getElementById('foto-rostro-input').value = '';
        document.getElementById('foto-preview').style.display = 'none';
        document.getElementById('btn-reiniciar').style.display = 'none';
        document.getElementById('btn-activar-cam').style.display = 'inline-block';
        document.getElementById('estado-foto').textContent = 'Sin foto registrada';
        document.getElementById('estado-foto').style.color = '#777';
    }

    function registrarRostroUsuario(usuario) {
        usuarioModal = usuario;
        document.getElementById('modal-rostro-user').textContent = usuario;
        document.getElementById('modal-rostro').style.display = 'flex';
        document.getElementById('estado-modal').textContent = 'Ponte frente a la camara';
        document.getElementById('estado-modal').style.color = '#2E75B6';
        var video = document.getElementById('camara-modal');
        navigator.mediaDevices.getUserMedia({ video: true })
        .then(function(stream) {
            streamModal = stream;
            video.srcObject = stream;
            video.play();
        })
        .catch(function(err) {
            alert('No se pudo acceder a la camara: ' + err.message);
            cerrarModalRostro();
        });
    }

    function capturarYGuardar() {
        var video = document.getElementById('camara-modal');
        var canvas = document.getElementById('canvas-modal');
        canvas.width = video.videoWidth;
        canvas.height = video.videoHeight;
        var ctx = canvas.getContext('2d');
        ctx.drawImage(video, 0, 0);
        var dataUrl = canvas.toDataURL('image/jpeg', 0.8);

        document.getElementById('estado-modal').textContent = 'Guardando rostro...';
        document.getElementById('estado-modal').style.color = '#2E75B6';

        fetch('/registrar-rostro', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ usuario: usuarioModal, foto: dataUrl })
        })
        .then(function(r) { return r.json(); })
        .then(function(data) {
            if (data.ok) {
                document.getElementById('estado-modal').textContent = 'Rostro registrado!';
                document.getElementById('estado-modal').style.color = '#28A745';
                setTimeout(function() { cerrarModalRostro(); location.reload(); }, 1000);
            } else {
                document.getElementById('estado-modal').textContent = data.error;
                document.getElementById('estado-modal').style.color = '#DC3545';
            }
        })
        .catch(function() {
            document.getElementById('estado-modal').textContent = 'Error de conexion';
            document.getElementById('estado-modal').style.color = '#DC3545';
        });
    }

    function cerrarModalRostro() {
        document.getElementById('modal-rostro').style.display = 'none';
        if (streamModal) {
            streamModal.getTracks().forEach(function(t) { t.stop(); });
            streamModal = null;
        }
    }
    </script>
    """ + HTML_FOOTER
)

PLANTILLA_INICIO = (
    HTML_HEAD + '<title>Control Financiero</title>' + ESTILOS + '</head><body>'
    + CABECERA +
    """
    <div class="resumen">
        <div class="tarjeta">
            <div class="numero" style="color: #28A745;">${{ '{:,.0f}'.format(total_ingresos) }}</div>
            <div class="etiqueta">Total Ingresos</div>
        </div>
        <div class="tarjeta">
            <div class="numero" style="color: #DC3545;">${{ '{:,.0f}'.format(total_egresos) }}</div>
            <div class="etiqueta">Total Egresos</div>
        </div>
        <div class="tarjeta">
            <div class="numero" style="color: {{ '#28A745' if saldo >= 0 else '#DC3545' }};">${{ '{:,.0f}'.format(saldo) }}</div>
            <div class="etiqueta">Saldo Actual</div>
        </div>
        <div class="tarjeta">
            <div class="numero">{{ movimientos|length }}</div>
            <div class="etiqueta">Movimientos</div>
        </div>
    </div>
    <div class="contenedor-tabla">
        <table>
            <thead>
                <tr>
                    <th>No.</th>
                    <th>Comprobante</th>
                    <th>Fecha</th>
                    <th>Concepto</th>
                    <th>Debito</th>
                    <th>Credito</th>
                    <th>Saldo</th>
                    <th>Medio</th>
                    <th>Referencia</th>
                    <th>Accion</th>
                </tr>
            </thead>
            <tbody>
                {% set saldo_acum = {'val': 0} %}
                {% for m in movimientos %}
                {% set _ = saldo_acum.update({'val': saldo_acum.val + m.credito - m.debito}) %}
                <tr>
                    <td>{{ loop.index }}</td>
                    <td>
                        {% if m.imagen %}
                        <div class="thumb-container" onclick="verImagen('/imagen/{{ m.imagen }}')">
                            <img src="/imagen/{{ m.imagen }}" alt="Comprobante {{ loop.index }}">
                        </div>
                        {% else %}
                        <div class="sin-imagen">--</div>
                        {% endif %}
                    </td>
                    <td>{{ m.fecha }}</td>
                    <td>{{ m.concepto }}</td>
                    <td class="valor" style="color: #DC3545;">{{ '${:,.0f}'.format(m.debito) if m.debito else '' }}</td>
                    <td class="valor" style="color: #28A745;">{{ '${:,.0f}'.format(m.credito) if m.credito else '' }}</td>
                    <td class="valor" style="color: {{ '#28A745' if saldo_acum.val >= 0 else '#DC3545' }};">${{ '{:,.0f}'.format(saldo_acum.val) }}</td>
                    <td>{{ m.medio }}</td>
                    <td>{{ m.referencia }}</td>
                    <td>
                        <form action="/eliminar/{{ loop.index0 }}" method="post"
                              onsubmit="return confirm('Eliminar movimiento #{{ loop.index }}?')">
                            <button type="submit" class="btn-eliminar-fila">Eliminar</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </tbody>
            <tfoot>
                <tr class="fila-total">
                    <td colspan="4">TOTAL</td>
                    <td class="valor">${{ '{:,.0f}'.format(total_egresos) }}</td>
                    <td class="valor">${{ '{:,.0f}'.format(total_ingresos) }}</td>
                    <td class="valor">${{ '{:,.0f}'.format(saldo) }}</td>
                    <td colspan="2">{{ movimientos|length }} movimientos</td>
                    <td></td>
                </tr>
            </tfoot>
        </table>
    </div>
    """
    + MODAL_SCRIPT + HTML_FOOTER
)

PLANTILLA_NUEVO = (
    HTML_HEAD + '<title>Agregar Movimiento</title>' + ESTILOS + '</head><body>'
    + CABECERA +
    """
    <div style="padding: 20px;">
        <div class="panel">
            <h2>Cargar Comprobante</h2>
            <form action="/procesar-imagen" method="post" enctype="multipart/form-data">
                <div class="form-campo" style="margin-bottom: 15px;">
                    <label>Seleccionar imagen del comprobante:</label>
                    <input type="file" name="imagen" accept="image/*" required
                           style="padding: 10px; border: 2px dashed #2E75B6; border-radius: 8px; width: 100%; cursor: pointer;">
                </div>
                <button type="submit" class="btn btn-azul">Leer imagen con OCR</button>
            </form>
            {% if datos %}
            <div style="padding: 10px; border-radius: 5px; margin: 10px 0; background: #D4EDDA; color: #155724; font-size: 13px; text-align: center;">
                Datos extraidos de la imagen. Revisa y confirma.
            </div>
            <form action="/agregar" method="post">
                <input type="hidden" name="imagen" value="{{ datos.imagen }}">
                <div class="form-grupo">
                    <div class="form-campo">
                        <label>Tipo:</label>
                        <select name="tipo" style="width:100%; padding:8px 12px; border:1px solid #CCC; border-radius:5px; font-size:14px;">
                            <option value="egreso" selected>Egreso (pago/gasto)</option>
                            <option value="ingreso">Ingreso (nomina/consignacion)</option>
                        </select>
                    </div>
                    <div class="form-campo">
                        <label>Fecha:</label>
                        <input type="text" name="fecha" value="{{ datos.fecha }}">
                    </div>
                    <div class="form-campo">
                        <label>Valor $:</label>
                        <input type="text" name="valor" value="{{ datos.valor }}">
                    </div>
                </div>
                <div class="form-grupo">
                    <div class="form-campo">
                        <label>Concepto:</label>
                        <input type="text" name="concepto" placeholder="Ej: Nomina, Arriendo, Servicios">
                    </div>
                    <div class="form-campo">
                        <label>Medio de Pago:</label>
                        <input type="text" name="medio" value="{{ datos.medio }}">
                    </div>
                    <div class="form-campo">
                        <label>Referencia:</label>
                        <input type="text" name="referencia" value="{{ datos.referencia }}">
                    </div>
                </div>
                <div class="form-grupo" style="grid-template-columns: 1fr;">
                    <div class="form-campo">
                        <label>Observacion:</label>
                        <input type="text" name="observacion" value="{{ datos.observacion }}">
                    </div>
                </div>
                <div style="text-align: right; margin-top: 15px;">
                    <button type="submit" class="btn btn-verde" style="font-size: 16px; padding: 12px 30px;">
                        CONFIRMAR Y REGISTRAR
                    </button>
                </div>
            </form>
            {% endif %}
        </div>
    </div>
    """ + HTML_FOOTER
)

PLANTILLA_CORREO = (
    HTML_HEAD + '<title>Enviar por Correo</title>' + ESTILOS + '</head><body>'
    + CABECERA +
    """
    <div style="padding: 20px;">
        <div class="panel">
            <h2>Enviar por Correo Electronico</h2>
            <p style="margin-bottom: 5px; color: #555;">
                <strong>Archivo:</strong> control_financiero.zip (Excel + comprobantes)<br>
                <strong>Total:</strong> ${{ '{:,.0f}'.format(total) }} - {{ num_pagos }} pagos<br>
                <strong>Remitente:</strong> {{ remitente }}
            </p>
            <h3 style="margin-top: 15px; font-size: 15px;">Destinatarios:</h3>
            <ul class="lista-correos">
                {% for correo in destinatarios %}
                <li>
                    <span>{{ correo }}</span>
                    <form action="/quitar-destinatario" method="post" style="display:inline;">
                        <input type="hidden" name="correo" value="{{ correo }}">
                        <button type="submit" class="btn-eliminar-fila">Quitar</button>
                    </form>
                </li>
                {% endfor %}
            </ul>
            <form action="/agregar-destinatario" method="post" class="fila-agregar">
                <input type="email" name="correo" placeholder="correo@ejemplo.com" required>
                <button type="submit" class="btn btn-azul">+ Agregar</button>
            </form>
            <div style="text-align: center; margin-top: 20px;">
                <form action="/enviar-correo" method="post"
                      onsubmit="this.querySelector('button').disabled=true; this.querySelector('button').textContent='Enviando...';">
                    <button type="submit" class="btn btn-rojo" style="font-size: 16px; padding: 12px 30px;">
                        ENVIAR CORREO
                    </button>
                </form>
            </div>
        </div>
    </div>
    """ + HTML_FOOTER
)


class Movimiento:
    def __init__(self, datos):
        self.fecha = datos.get('fecha', '')
        self.valor = datos.get('valor', 0)
        self.tipo = datos.get('tipo', 'egreso')
        self.concepto = datos.get('concepto', '')
        self.medio = datos.get('medio', '')
        self.referencia = datos.get('referencia', '')
        self.observacion = datos.get('observacion', '')
        self.imagen = datos.get('imagen', '')

    @property
    def debito(self):
        return self.valor if self.tipo == 'egreso' else 0

    @property
    def credito(self):
        return self.valor if self.tipo == 'ingreso' else 0


def get_fechas(pagos_raw):
    if pagos_raw:
        return pagos_raw[0].get('fecha', ''), pagos_raw[-1].get('fecha', '')
    return '', ''


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip()
        clave = request.form.get('clave', '')
        usuarios = cargar_usuarios()
        if usuario in usuarios and usuarios[usuario]['clave'] == hash_clave(clave):
            session['usuario'] = usuario
            session['nombre'] = usuarios[usuario]['nombre']
            flash(f'Bienvenido(a), {usuarios[usuario]["nombre"]}.', 'exito')
            return redirect(url_for('inicio'))
        flash('Usuario o contrasena incorrectos.', 'error')
        return redirect(url_for('login'))
    return render_template_string(PLANTILLA_LOGIN)


@app.route('/login-facial', methods=['POST'])
def login_facial():
    datos = request.get_json()
    foto = datos.get('foto', '')
    if not foto:
        return jsonify(ok=False, error='No se recibio la foto'), 400
    usuario = verificar_rostro(foto)
    if usuario:
        usuarios = cargar_usuarios()
        session['usuario'] = usuario
        session['nombre'] = usuarios[usuario]['nombre']
        return jsonify(ok=True, nombre=usuarios[usuario]['nombre'])
    return jsonify(ok=False, error='Rostro no reconocido')


@app.route('/logout')
def logout():
    session.clear()
    flash('Sesion cerrada correctamente.', 'exito')
    return redirect(url_for('login'))


@app.route('/usuarios')
@admin_requerido
def usuarios():
    usuarios = cargar_usuarios()
    return render_template_string(PLANTILLA_USUARIOS, usuarios=usuarios,
                                  fecha_desde='', fecha_hasta='')


@app.route('/crear-usuario', methods=['POST'])
@admin_requerido
def crear_usuario():
    usuario = request.form.get('usuario', '').strip().lower()
    nombre = request.form.get('nombre', '').strip()
    clave = request.form.get('clave', '')
    if not usuario or not nombre or not clave:
        flash('Todos los campos son obligatorios.', 'error')
        return redirect(url_for('usuarios'))
    if len(clave) < 4:
        flash('La contrasena debe tener al menos 4 caracteres.', 'error')
        return redirect(url_for('usuarios'))
    usuarios_dict = cargar_usuarios()
    if usuario in usuarios_dict:
        flash(f'El usuario "{usuario}" ya existe.', 'error')
        return redirect(url_for('usuarios'))
    usuarios_dict[usuario] = {
        'nombre': nombre,
        'clave': hash_clave(clave)
    }
    guardar_usuarios(usuarios_dict)
    foto_rostro = request.form.get('foto_rostro', '')
    msg_rostro = ''
    if foto_rostro:
        if registrar_rostro(usuario, foto_rostro):
            msg_rostro = ' con reconocimiento facial activado'
        else:
            msg_rostro = ' (no se detecto un rostro en la foto, se puede registrar despues)'
    flash(f'Usuario "{usuario}" creado exitosamente{msg_rostro}.', 'exito')
    return redirect(url_for('usuarios'))


@app.route('/registrar-rostro', methods=['POST'])
@admin_requerido
def registrar_rostro_ruta():
    datos = request.get_json()
    usuario = datos.get('usuario', '')
    foto = datos.get('foto', '')
    if not usuario or not foto:
        return jsonify(ok=False, error='Datos incompletos')
    usuarios_dict = cargar_usuarios()
    if usuario not in usuarios_dict:
        return jsonify(ok=False, error='Usuario no existe')
    if registrar_rostro(usuario, foto):
        return jsonify(ok=True)
    return jsonify(ok=False, error='No se detecto un rostro en la foto. Intenta de nuevo.')


@app.route('/eliminar-usuario', methods=['POST'])
@admin_requerido
def eliminar_usuario():
    usuario = request.form.get('usuario', '').strip()
    if usuario == 'admin':
        flash('No se puede eliminar al administrador.', 'error')
        return redirect(url_for('usuarios'))
    usuarios_dict = cargar_usuarios()
    if usuario in usuarios_dict:
        del usuarios_dict[usuario]
        guardar_usuarios(usuarios_dict)
        flash(f'Usuario "{usuario}" eliminado.', 'exito')
    return redirect(url_for('usuarios'))


@app.route('/')
@login_requerido
def inicio():
    datos_raw = cargar_pagos()
    movimientos = [Movimiento(p) for p in datos_raw]
    total_ingresos = sum(m.credito for m in movimientos)
    total_egresos = sum(m.debito for m in movimientos)
    saldo = total_ingresos - total_egresos
    fd, fh = get_fechas(datos_raw)
    return render_template_string(PLANTILLA_INICIO, movimientos=movimientos,
                                  total_ingresos=total_ingresos, total_egresos=total_egresos,
                                  saldo=saldo, fecha_desde=fd, fecha_hasta=fh)


@app.route('/imagen/<nombre>')
@login_requerido
def servir_imagen(nombre):
    return send_from_directory(RUTA_APP, nombre)


@app.route('/foto-usuario/<usuario>')
@login_requerido
def foto_usuario(usuario):
    ruta_foto = os.path.join(CARPETA_ROSTROS, f'{usuario}.jpg')
    if os.path.exists(ruta_foto):
        return send_from_directory(CARPETA_ROSTROS, f'{usuario}.jpg')
    return '', 404


@app.route('/nuevo')
@login_requerido
def nuevo():
    return render_template_string(PLANTILLA_NUEVO, datos=None,
                                  fecha_desde='', fecha_hasta='')


@app.route('/procesar-imagen', methods=['POST'])
@login_requerido
def procesar_imagen():
    if 'imagen' not in request.files:
        flash('No se selecciono ninguna imagen.', 'error')
        return redirect(url_for('nuevo'))

    archivo = request.files['imagen']
    if archivo.filename == '':
        flash('No se selecciono ninguna imagen.', 'error')
        return redirect(url_for('nuevo'))

    nombre = archivo.filename
    ruta_destino = os.path.join(RUTA_APP, nombre)
    archivo.save(ruta_destino)

    try:
        datos = extraer_datos_imagen(ruta_destino)
        datos['imagen'] = nombre
    except Exception as e:
        flash(f'Error al leer la imagen con OCR: {e}', 'error')
        return redirect(url_for('nuevo'))

    pagos = cargar_pagos()
    for p in pagos:
        if p.get('imagen') == nombre:
            flash(f'Esta imagen ya fue registrada (Fecha: {p["fecha"]}, Valor: ${p["valor"]:,.0f}).', 'error')
            return redirect(url_for('nuevo'))
        if datos['referencia'] and p.get('referencia') == datos['referencia']:
            flash(f'Ya existe un pago con la misma referencia: {datos["referencia"]}.', 'error')
            return redirect(url_for('nuevo'))

    return render_template_string(PLANTILLA_NUEVO, datos=datos,
                                  fecha_desde='', fecha_hasta='')


@app.route('/agregar', methods=['POST'])
@login_requerido
def agregar():
    fecha = request.form.get('fecha', '').strip()
    valor_str = request.form.get('valor', '0').strip().replace('.', '').replace(',', '').replace('$', '')
    tipo = request.form.get('tipo', 'egreso').strip()
    concepto = request.form.get('concepto', '').strip()
    medio = request.form.get('medio', '').strip()
    referencia = request.form.get('referencia', '').strip()
    observacion = request.form.get('observacion', '').strip()
    imagen = request.form.get('imagen', '').strip()

    if not fecha or not valor_str:
        flash('Fecha y Valor son obligatorios.', 'error')
        return redirect(url_for('nuevo'))

    try:
        valor = int(valor_str)
    except ValueError:
        flash('El valor debe ser un numero.', 'error')
        return redirect(url_for('nuevo'))

    movimiento = {
        'fecha': fecha, 'valor': valor, 'tipo': tipo, 'concepto': concepto,
        'medio': medio, 'referencia': referencia, 'observacion': observacion,
        'imagen': imagen
    }

    pagos = cargar_pagos()
    pagos.append(movimiento)
    guardar_pagos(pagos)

    etiqueta = 'Ingreso' if tipo == 'ingreso' else 'Egreso'
    flash(f'{etiqueta} #{len(pagos)} por ${valor:,.0f} registrado correctamente.', 'exito')
    return redirect(url_for('inicio'))


@app.route('/eliminar/<int:indice>', methods=['POST'])
@login_requerido
def eliminar(indice):
    pagos = cargar_pagos()
    if 0 <= indice < len(pagos):
        pago = pagos.pop(indice)
        guardar_pagos(pagos)
        flash(f'Movimiento #{indice+1} ({pago["fecha"]} - ${pago["valor"]:,.0f}) eliminado.', 'exito')
    return redirect(url_for('inicio'))


@app.route('/descargar-excel')
@login_requerido
def descargar_excel():
    pagos = cargar_pagos()
    if not pagos:
        flash('No hay movimientos registrados.', 'error')
        return redirect(url_for('inicio'))
    generar_excel(pagos)
    return send_file(ARCHIVO_EXCEL, as_attachment=True, download_name='control_financiero.xlsx')


@app.route('/correo')
@login_requerido
def correo():
    pagos = cargar_pagos()
    total_ingresos = sum(p['valor'] for p in pagos if p.get('tipo', 'egreso') == 'ingreso')
    total_egresos = sum(p['valor'] for p in pagos if p.get('tipo', 'egreso') == 'egreso')
    saldo = total_ingresos - total_egresos
    destinatarios = cargar_destinatarios()
    return render_template_string(
        PLANTILLA_CORREO, total=saldo, num_pagos=len(pagos),
        remitente=CORREO_REMITENTE, destinatarios=destinatarios,
        fecha_desde='', fecha_hasta='')


@app.route('/agregar-destinatario', methods=['POST'])
@login_requerido
def agregar_destinatario():
    correo_nuevo = request.form.get('correo', '').strip().lower()
    if not correo_nuevo or '@' not in correo_nuevo:
        flash('Escribe un correo valido.', 'error')
        return redirect(url_for('correo'))

    destinatarios = cargar_destinatarios()
    if correo_nuevo in destinatarios:
        flash('Ese correo ya esta en la lista.', 'info')
        return redirect(url_for('correo'))

    destinatarios.append(correo_nuevo)
    guardar_destinatarios(destinatarios)
    flash(f'Se agrego {correo_nuevo}.', 'exito')
    return redirect(url_for('correo'))


@app.route('/quitar-destinatario', methods=['POST'])
@login_requerido
def quitar_destinatario():
    correo_quitar = request.form.get('correo', '')
    destinatarios = cargar_destinatarios()
    if correo_quitar in destinatarios:
        destinatarios.remove(correo_quitar)
        guardar_destinatarios(destinatarios)
        flash(f'Se quito {correo_quitar}.', 'exito')
    return redirect(url_for('correo'))


@app.route('/enviar-correo', methods=['POST'])
@login_requerido
def enviar_correo():
    pagos = cargar_pagos()
    if not pagos:
        flash('No hay movimientos registrados.', 'error')
        return redirect(url_for('correo'))

    destinatarios = cargar_destinatarios()
    if not destinatarios:
        flash('Agrega al menos un destinatario.', 'error')
        return redirect(url_for('correo'))

    saldo = generar_excel(pagos)

    try:
        msg = MIMEMultipart()
        msg['From'] = CORREO_REMITENTE
        msg['To'] = ', '.join(destinatarios)
        msg['Subject'] = f'Control Financiero - {len(pagos)} movimientos - Saldo: ${saldo:,.0f}'

        cuerpo = (
            f"Control Financiero\n"
            f"{'=' * 50}\n\n"
            f"Total movimientos: {len(pagos)}\n"
            f"Saldo actual: ${saldo:,.0f}\n"
            f"Fecha de envio: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
            f"Se adjunta un ZIP con el Excel y todos los comprobantes.\n"
            f"Descomprime el ZIP en una carpeta y abre el Excel desde ahi.\n"
            f"Los enlaces 'Ver comprobante' abren las imagenes directamente.\n"
        )
        msg.attach(MIMEText(cuerpo, 'plain'))

        archivo_zip = os.path.join(RUTA_APP, 'control_financiero.zip')
        with zipfile.ZipFile(archivo_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(ARCHIVO_EXCEL, 'control_financiero.xlsx')
            for pago in pagos:
                if pago.get('imagen'):
                    ruta_img = os.path.join(RUTA_APP, pago['imagen'])
                    if os.path.exists(ruta_img):
                        zf.write(ruta_img, pago['imagen'])

        with open(archivo_zip, 'rb') as f:
            adjunto_zip = MIMEBase('application', 'octet-stream')
            adjunto_zip.set_payload(f.read())
            encoders.encode_base64(adjunto_zip)
            adjunto_zip.add_header('Content-Disposition',
                                   'attachment', filename='control_financiero.zip')
            msg.attach(adjunto_zip)

        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(CORREO_REMITENTE, CORREO_CLAVE_APP)
        servidor.sendmail(CORREO_REMITENTE, destinatarios, msg.as_string())
        servidor.quit()

        destinos = ', '.join(destinatarios)
        flash(f'Correo enviado exitosamente a: {destinos}', 'exito')

    except Exception as e:
        flash(f'Error al enviar correo: {e}', 'error')

    return redirect(url_for('correo'))


if __name__ == '__main__':
    print("=" * 50)
    print("  Control Financiero - Servidor iniciado")
    print("  Abre en tu navegador: http://localhost:5050")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5050, debug=True)
