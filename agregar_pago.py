#!/usr/bin/env python3
"""
Script para agregar un nuevo pago al archivo pagos_liliana.xlsx
Uso: python3 agregar_pago.py
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

RUTA = os.path.dirname(os.path.abspath(__file__))
ARCHIVO = os.path.join(RUTA, 'pagos_liliana.xlsx')

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
link_font = Font(color="0563C1", underline="single")

print("=" * 50)
print("  AGREGAR NUEVO PAGO - Liliana Moreno")
print("=" * 50)
print()

# Pedir datos
fecha = input("Fecha del pago (dd/mm/aaaa): ").strip()
valor = input("Valor del pago (solo numero, ej: 200000): ").strip()
valor = int(valor.replace('.', '').replace(',', '').replace('$', ''))
medio = input("Medio de pago (ej: Nequi, Banco de Bogota): ").strip()
referencia = input("Referencia o No. autorizacion: ").strip()
observacion = input("Observacion (opcional, Enter para omitir): ").strip()
imagen = input("Nombre del archivo de imagen (ej: 14_pago_200000.jpeg): ").strip()

# Abrir Excel existente
wb = openpyxl.load_workbook(ARCHIVO)
ws = wb.active

# Encontrar la fila de TOTAL para insertar antes
total_row = None
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == 'TOTAL':
        total_row = row
        break

if total_row is None:
    print("Error: no se encontro la fila TOTAL en el Excel")
    exit(1)

# Obtener numero de cuota y acumulado anterior
cuota_anterior = ws.cell(row=total_row - 1, column=1).value
acumulado_anterior = ws.cell(row=total_row - 1, column=4).value or 0
nueva_cuota = cuota_anterior + 1 if isinstance(cuota_anterior, int) else 1
nuevo_acumulado = acumulado_anterior + valor

# Insertar fila antes de TOTAL
ws.insert_rows(total_row)
nueva_fila = total_row  # la fila insertada queda donde estaba TOTAL

datos = [nueva_cuota, fecha, valor, nuevo_acumulado, medio, referencia, observacion]
for col_idx, val in enumerate(datos, 1):
    cell = ws.cell(row=nueva_fila, column=col_idx, value=val)
    cell.border = border
    if col_idx in (3, 4):
        cell.number_format = '$#,##0'
        cell.alignment = Alignment(horizontal='right')
    elif col_idx in (1, 2):
        cell.alignment = Alignment(horizontal='center')

# Hipervinculo a la imagen
if imagen:
    cell_link = ws.cell(row=nueva_fila, column=8, value='Ver comprobante')
    cell_link.hyperlink = './' + imagen
    cell_link.font = link_font
    cell_link.alignment = Alignment(horizontal='center')
    cell_link.border = border

# Actualizar fila TOTAL (ahora esta una fila mas abajo)
total_row_new = total_row + 1
ws.cell(row=total_row_new, column=3, value=nuevo_acumulado)
ws.cell(row=total_row_new, column=3).number_format = '$#,##0'
ws.cell(row=total_row_new, column=3).font = Font(bold=True, size=11)
ws.cell(row=total_row_new, column=3).alignment = Alignment(horizontal='right')
ws.cell(row=total_row_new, column=7, value=f'{nueva_cuota} pagos realizados')

wb.save(ARCHIVO)
print()
print(f"Pago #{nueva_cuota} agregado exitosamente!")
print(f"Valor: ${valor:,.0f}")
print(f"Acumulado: ${nuevo_acumulado:,.0f}")
print(f"Archivo actualizado: {ARCHIVO}")
