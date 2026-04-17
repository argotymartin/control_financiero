#!/usr/bin/env python3
"""
App Web - Control Financiero (Version Supabase)
Misma funcionalidad que web_pagos.py pero usando Supabase como backend.
Plantillas en templates/ con herencia Jinja2.
Autenticacion: Supabase Auth (OAuth/Email)
"""

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    session,
    jsonify,
)
from functools import wraps
from datetime import datetime
import json
import os
import re
import smtplib
import subprocess
import tempfile
import zipfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from io import BytesIO

import base64
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import face_recognition
import numpy as np
from supabase import create_client

# ==================== CONFIGURACION SUPABASE ====================
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")

# Compatibilidad: acepta SUPABASE_KEY (viejo) o SUPABASE_ANON_KEY (nuevo)
SUPABASE_ANON_KEY = os.environ.get(
    "SUPABASE_ANON_KEY", os.environ.get("SUPABASE_KEY", "")
)
SUPABASE_SERVICE_KEY = os.environ.get(
    "SUPABASE_SERVICE_KEY",
    os.environ.get("SUPABASE_SERVICE_ROLE_KEY", os.environ.get("SUPABASE_KEY", "")),
)

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    print("=" * 60)
    print("  ERROR: Configura las variables de entorno:")
    print("    export SUPABASE_URL='https://tu-proyecto.supabase.co'")
    print("    export SUPABASE_ANON_KEY='tu-anon-key'")
    print("    export SUPABASE_SERVICE_KEY='tu-service-role-key'")
    print("=" * 60)
    exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

supabase_admin = (
    create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY) if SUPABASE_SERVICE_KEY else None
)

CORREO_REMITENTE = os.environ.get("CORREO_REMITENTE", "argoty.martin@gmail.com")
CORREO_CLAVE_APP = os.environ.get("CORREO_CLAVE_APP", "")

ADMIN_USUARIOS = os.environ.get("ADMIN_USUARIOS", "admin,admin@test.com").split(",")

MESES = {
    "enero": "01",
    "febrero": "02",
    "marzo": "03",
    "abril": "04",
    "mayo": "05",
    "junio": "06",
    "julio": "07",
    "agosto": "08",
    "septiembre": "09",
    "octubre": "10",
    "noviembre": "11",
    "diciembre": "12",
}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pagos_liliana_supabase_2024")


# ==================== CONTEXT PROCESSOR ====================
# Inyecta fecha_desde y fecha_hasta en todas las plantillas automaticamente


@app.context_processor
def inyectar_fechas():
    datos = {"fecha_desde": "", "fecha_hasta": ""}
    if "email" in session:
        datos_raw = cargar_pagos()
        fd, fh = get_fechas(datos_raw)
        datos["fecha_desde"] = fd
        datos["fecha_hasta"] = fh
        datos["es_admin"] = es_admin()
        datos["nombre_usuario"] = session.get("nombre", "")
        datos["email_usuario"] = session.get("email", "")
    return datos


# ==================== FUNCIONES DE BASE DE DATOS ====================


def cargar_pagos():
    client = supabase_admin if supabase_admin else supabase
    resp = client.table("pagos").select("*").order("id").execute()
    return resp.data


def guardar_pago(pago):
    if supabase_admin:
        supabase_admin.table("pagos").insert(pago).execute()
    else:
        supabase.table("pagos").insert(pago).execute()


def eliminar_pago(pago_id):
    client = supabase_admin if supabase_admin else supabase
    client.table("pagos").delete().eq("id", pago_id).execute()


def cargar_destinatarios():
    resp = supabase.table("destinatarios").select("correo").execute()
    return [d["correo"] for d in resp.data]


def agregar_destinatario_db(correo):
    supabase.table("destinatarios").insert({"correo": correo}).execute()


def quitar_destinatario_db(correo):
    supabase.table("destinatarios").delete().eq("correo", correo).execute()


def obtener_no_vistos(usuario):
    pagos = cargar_pagos()
    pagos_con_imagen = [p for p in pagos if p.get("imagen")]
    if not pagos_con_imagen:
        return 0
    client = supabase_admin if supabase_admin else supabase
    vistos = (
        client.table("pagos_vistos")
        .select("pago_id")
        .eq("usuario", usuario)
        .execute()
    )
    ids_vistos = {v["pago_id"] for v in vistos.data}
    return sum(1 for p in pagos_con_imagen if p["id"] not in ids_vistos)


def marcar_visto(usuario, pago_id):
    client = supabase_admin if supabase_admin else supabase
    client.table("pagos_vistos").upsert(
        {"usuario": usuario, "pago_id": pago_id}
    ).execute()


# ==================== SUPABASE STORAGE ====================


def subir_imagen_storage(bucket, nombre, datos_bytes, content_type="image/jpeg"):
    client = supabase_admin if supabase_admin else supabase
    client.storage.from_(bucket).upload(
        nombre,
        datos_bytes,
        file_options={"content-type": content_type, "upsert": "true"},
    )


def obtener_url_publica(bucket, nombre):
    return supabase.storage.from_(bucket).get_public_url(nombre)


# ==================== PUSH NOTIFICATIONS ====================

VAPID_PUBLIC_KEY = "BCzFQUfKsgTTvEHJAf0nab14ePgbp_A42PqxSbbtEradsTO8vkssn_fEszWLp500nV35STGjezTkqUwTeBQtegk"
VAPID_PRIVATE_KEY = "-EYrty98AZU6wPAu2P4NvPBNkRyvi3-PFru4VwRros8"


def guardar_suscripcion_push(usuario, endpoint, auth, p256dh):
    """Guardar suscripción de push en Supabase"""
    try:
        client = supabase_admin if supabase_admin else supabase
        client.table("push_subscriptions_control_financiero").upsert(
            {
                "usuario": usuario,
                "endpoint": endpoint,
                "auth": auth,
                "p256dh": p256dh,
            },
            on_conflict="endpoint"
        ).execute()
        return True
    except Exception as e:
        print(f"Error guardando suscripción push: {e}")
        return False


def obtener_suscripciones_push():
    """Obtener todas las suscripciones de push"""
    try:
        resp = supabase.table("push_subscriptions_control_financiero").select("*").execute()
        return resp.data
    except Exception as e:
        print(f"Error obteniendo suscripciones push: {e}")
        return []


def enviar_notificacion_push(titulo, cuerpo):
    """Enviar notificación push a todos los usuarios suscritos"""
    from pywebpush import webpush
    from urllib.parse import urlparse

    suscripciones = obtener_suscripciones_push()

    for sub in suscripciones:
        try:
            datos = {
                "title": titulo,
                "body": cuerpo
            }

            # Extraer host del endpoint
            parsed_url = urlparse(sub["endpoint"])
            aud = f"{parsed_url.scheme}://{parsed_url.netloc}"

            webpush(
                subscription_info={
                    "endpoint": sub["endpoint"],
                    "keys": {
                        "auth": sub["auth"],
                        "p256dh": sub["p256dh"]
                    }
                },
                data=json.dumps(datos),
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={
                    "sub": "mailto:contact@example.com",
                    "aud": aud
                }
            )
            print(f"Notificación enviada a {sub['usuario']}")
        except Exception as e:
            print(f"Error enviando notificación a {sub['usuario']}: {e}")


def descargar_imagen_storage(bucket, nombre):
    return supabase.storage.from_(bucket).download(nombre)


# ==================== AUTENTICACION SUPABASE AUTH ====================


def obtener_usuario_actual():
    """Obtiene el usuario actual desde la sesion de Flask."""
    return session.get("usuario"), session.get("email"), session.get("nombre")


def es_admin():
    """Verifica si el usuario actual es admin."""
    email = session.get("email", "")
    return email in ADMIN_USUARIOS


def login_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        usuario, email, nombre = obtener_usuario_actual()
        if not email:
            flash("Debes iniciar sesion para acceder.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorador


def admin_requerido(f):
    @wraps(f)
    @login_requerido
    def decorador(*args, **kwargs):
        if not es_admin():
            flash("Solo el administrador puede acceder a esta funcion.", "error")
            return redirect(url_for("inicio"))
        return f(*args, **kwargs)

    return decorador


# ==================== RUTAS DE AUTH ====================


@app.route("/auth/login", methods=["GET", "POST"])
def auth_login():
    return render_template("paginas/login.html")


@app.route("/auth/ingresar", methods=["POST"])
def auth_ingresar():
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")

    if not email or not password:
        flash("Email y contrasena son requeridos.", "error")
        return redirect(url_for("auth_login"))

    try:
        resp = supabase.auth.sign_in_with_password(
            {"email": email, "password": password}
        )

        if resp.user:
            session["usuario"] = resp.user.id
            session["email"] = resp.user.email
            session["nombre"] = resp.user.email.split("@")[0]
            session["access_token"] = resp.session.access_token
            session["refresh_token"] = resp.session.refresh_token

            flash(f"Bienvenido(a), {session['nombre']}.", "exito")
            return redirect(url_for("inicio"))

    except Exception as e:
        flash(f"Error al iniciar sesion: {str(e)}", "error")

    return redirect(url_for("auth_login"))


@app.route("/auth/registro", methods=["GET", "POST"])
def auth_registro():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        nombre = request.form.get("nombre", "").strip()

        if not email or not password or not nombre:
            flash("Todos los campos son requeridos.", "error")
            return redirect(url_for("auth_registro"))

        if len(password) < 6:
            flash("La contrasena debe tener al menos 6 caracteres.", "error")
            return redirect(url_for("auth_registro"))

        try:
            resp = supabase.auth.sign_up(
                {
                    "email": email,
                    "password": password,
                    "options": {"data": {"nombre": nombre}},
                }
            )

            if resp.user:
                flash("Cuenta creada. Revisa tu correo para verificar.", "exito")
            else:
                flash("Error al crear la cuenta.", "error")

        except Exception as e:
            flash(f"Error al registrar: {str(e)}", "error")

    return render_template("paginas/registro.html")


@app.route("/auth/google")
def auth_google():
    redirect_uri = url_for("auth_callback", _external=True)
    return supabase.auth.sign_in_with_oauth(
        {"provider": "google", "options": {"redirect_to": redirect_uri}}
    )


@app.route("/auth/callback")
def auth_callback():
    code = request.args.get("code")
    if code:
        try:
            resp = supabase.auth.exchange_code_for_session(code)
            if resp.user:
                session["usuario"] = resp.user.id
                session["email"] = resp.user.email
                session["nombre"] = resp.user.user_metadata.get(
                    "full_name", resp.user.email.split("@")[0]
                )
                session["access_token"] = resp.session.access_token
                session["refresh_token"] = resp.session.refresh_token
                flash(f"Bienvenido(a), {session['nombre']}.", "exito")
                return redirect(url_for("inicio"))
        except Exception as e:
            flash(f"Error en autenticacion: {str(e)}", "error")
    return redirect(url_for("auth_login"))


@app.route("/auth/logout")
def auth_logout():
    try:
        supabase.auth.sign_out()
    except:
        pass
    session.clear()
    flash("Sesion cerrada correctamente.", "exito")
    return redirect(url_for("auth_login"))


@app.route("/auth/recuperar", methods=["GET", "POST"])
def auth_recuperar():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if not email:
            flash("Ingresa tu correo electronico.", "error")
            return redirect(url_for("auth_recuperar"))

        try:
            supabase.auth.reset_password_email(email)
            flash("Se ha enviado un enlace de recuperacion a tu correo.", "exito")
        except Exception as e:
            flash(f"Error: {str(e)}", "error")

    return render_template("paginas/recuperar.html")


@app.route("/auth/actualizar-password", methods=["GET", "POST"])
def auth_actualizar_password():
    if request.method == "POST":
        password = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        if not password or len(password) < 6:
            flash("Contrasena minima de 6 caracteres.", "error")
            return redirect(url_for("auth_actualizar_password"))

        if password != password2:
            flash("Las contrasenas no coinciden.", "error")
            return redirect(url_for("auth_actualizar_password"))

        access_token = session.get("access_token")
        if not access_token:
            flash("Sesion expirada.", "error")
            return redirect(url_for("auth_login"))

        try:
            supabase.auth.update_user({"password": password})
            flash("Contrasena actualizada correctamente.", "exito")
            return redirect(url_for("inicio"))
        except Exception as e:
            flash(f"Error: {str(e)}", "error")

    return render_template("paginas/actualizar_password.html")


# ==================== FACIAL RECOGNITION ====================

def guardar_imagen_base64(imagen_base64, ruta):
    """Guardar imagen base64 a archivo"""
    if imagen_base64.startswith('data:image'):
        imagen_base64 = imagen_base64.split(',')[1]
    datos_imagen = base64.b64decode(imagen_base64)
    with open(ruta, 'wb') as f:
        f.write(datos_imagen)


def obtener_encoding_facial(ruta_imagen):
    """Obtener encoding facial de una imagen"""
    try:
        imagen = face_recognition.load_image_file(ruta_imagen)
        encodings = face_recognition.face_encodings(imagen)
        if encodings:
            return encodings[0].tolist()
        return None
    except Exception as e:
        print(f"Error obteniendo encoding facial: {e}")
        return None


def verificar_rostro_facial(imagen_base64):
    """Verificar rostro en BD Supabase"""
    ruta_temp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False).name
    try:
        guardar_imagen_base64(imagen_base64, ruta_temp)
        encoding_login = obtener_encoding_facial(ruta_temp)

        if encoding_login is None:
            return None

        # Obtener todos los rostros de la BD
        client = supabase_admin if supabase_admin else supabase
        resp = client.table("rostros_usuarios").select("email, encoding").execute()
        rostros = resp.data

        for rostro in rostros:
            try:
                # Si encoding es string JSON, parsear
                encoding_guardado = rostro['encoding']
                if isinstance(encoding_guardado, str):
                    encoding_guardado = json.loads(encoding_guardado)

                encoding_guardado = np.array(encoding_guardado, dtype=np.float64)
                encoding_login_array = np.array(encoding_login, dtype=np.float64)

                resultado = face_recognition.compare_faces([encoding_guardado], encoding_login_array, tolerance=0.6)
                if resultado[0]:
                    return rostro['email']
            except Exception as e:
                print(f"Error comparando rostros: {e}")
                continue

        return None
    finally:
        if os.path.exists(ruta_temp):
            os.remove(ruta_temp)


@app.route('/auth/login-facial', methods=['POST'])
def auth_login_facial():
    """Login con reconocimiento facial"""
    try:
        datos = request.get_json()
        foto = datos.get('foto', '')

        if not foto:
            return jsonify(ok=False, error='No se recibio la foto'), 400

        email = verificar_rostro_facial(foto)

        if not email:
            return jsonify(ok=False, error='Rostro no reconocido'), 401

        # Guardar en sesion (sin necesidad de password)
        session['email'] = email
        session['nombre'] = email.split('@')[0]
        return jsonify(ok=True, nombre=email.split('@')[0])

    except Exception as e:
        print(f"Error en login facial: {e}")
        return jsonify(ok=False, error='Error en verificacion'), 500


@app.route('/registrar-rostro')
@login_requerido
def registrar_rostro():
    """Página para registrar rostro"""
    return render_template('paginas/registrar_rostro.html')


@app.route('/api/registrar-rostro', methods=['POST'])
@login_requerido
def api_registrar_rostro():
    """Guardar rostro del usuario actual"""
    try:
        datos = request.get_json()
        foto = datos.get('foto', '')

        if not foto:
            return jsonify(ok=False, error='No se recibio la foto'), 400

        email = session.get('email')
        if not email:
            return jsonify(ok=False, error='No autenticado'), 401

        encoding = obtener_encoding_facial_base64(foto)
        if not encoding:
            return jsonify(ok=False, error='No se detecto rostro en la foto'), 400

        # Guardar en Supabase
        client = supabase_admin if supabase_admin else supabase
        client.table('rostros_usuarios').upsert({
            'email': email,
            'encoding': encoding
        }, on_conflict='email').execute()

        try:
            subir_foto_rostro(email, foto)
        except Exception as e:
            print(f"Error subiendo JPG rostro: {e}")

        return jsonify(ok=True, mensaje='Rostro registrado correctamente')

    except Exception as e:
        print(f"Error registrando rostro: {e}")
        return jsonify(ok=False, error=str(e)), 500


def obtener_encoding_facial_base64(imagen_base64):
    """Obtener encoding facial de imagen en base64"""
    ruta_temp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False).name
    try:
        guardar_imagen_base64(imagen_base64, ruta_temp)
        encoding = obtener_encoding_facial(ruta_temp)
        return encoding
    finally:
        if os.path.exists(ruta_temp):
            os.remove(ruta_temp)


def subir_foto_rostro(email, foto_base64):
    """Sube el JPG de rostro al bucket 'rostros' con nombre {email}.jpg"""
    if "," in foto_base64:
        foto_base64 = foto_base64.split(",", 1)[1]
    datos = base64.b64decode(foto_base64)
    nombre = f"{email}.jpg"
    client = supabase_admin if supabase_admin else supabase
    try:
        client.storage.from_("rostros").remove([nombre])
    except Exception:
        pass
    client.storage.from_("rostros").upload(
        nombre, datos, {"content-type": "image/jpeg", "upsert": "true"}
    )


@app.route('/usuarios')
@login_requerido
def usuarios():
    """Página de gestión de usuarios"""
    if session.get('email') != 'admin@test.com':
        flash('Solo admin puede acceder.', 'error')
        return redirect(url_for('inicio'))
    return render_template('paginas/usuarios.html')


@app.route('/api/usuarios')
@login_requerido
def api_usuarios():
    """API para obtener lista de usuarios"""
    try:
        if session.get('email') != 'admin@test.com':
            return jsonify(ok=False, error='No autorizado'), 403

        client = supabase_admin if supabase_admin else supabase

        # Traer TODOS los usuarios paginando
        todos = []
        pagina = 1
        while True:
            try:
                resp = client.auth.admin.list_users(page=pagina, per_page=1000)
            except TypeError:
                resp = client.auth.admin.list_users()
            lote = resp.users if hasattr(resp, 'users') else (resp or [])
            if not lote:
                break
            todos.extend(lote)
            if len(lote) < 1000:
                break
            pagina += 1

        usuarios_list = []
        for user in todos:
            try:
                rostro_resp = client.table('rostros_usuarios').select('*').eq('email', user.email).execute()
                tiene_rostro = len(rostro_resp.data) > 0
            except Exception:
                tiene_rostro = False

            usuarios_list.append({
                'email': user.email,
                'nombre': user.email.split('@')[0],
                'tiene_rostro': tiene_rostro
            })

        return jsonify(ok=True, usuarios=usuarios_list)
    except Exception as e:
        print(f"Error obteniendo usuarios: {e}")
        return jsonify(ok=False, error=str(e)), 500


@app.route('/api/crear-usuario', methods=['POST'])
@login_requerido
def api_crear_usuario():
    """Crear nuevo usuario en Supabase Auth"""
    try:
        if session.get('email') != 'admin@test.com':
            return jsonify(ok=False, error='No autorizado'), 403

        datos = request.get_json()
        email = datos.get('email', '').strip().lower()
        password = datos.get('password', '')
        nombre = datos.get('nombre', '')

        if not email or not password or len(password) < 6:
            return jsonify(ok=False, error='Datos incompletos o password < 6 caracteres'), 400

        # Crear en Supabase Auth
        client = supabase_admin if supabase_admin else supabase
        try:
            resp = client.auth.admin.create_user(
                {"email": email, "password": password, "email_confirm": True}
            )
            return jsonify(ok=True, mensaje='Usuario creado correctamente')
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 400

    except Exception as e:
        print(f"Error creando usuario: {e}")
        return jsonify(ok=False, error=str(e)), 500


@app.route('/api/registrar-rostro-admin', methods=['POST'])
@login_requerido
def api_registrar_rostro_admin():
    """Registrar rostro de usuario (admin)"""
    try:
        if session.get('email') != 'admin@test.com':
            return jsonify(ok=False, error='No autorizado'), 403

        datos = request.get_json()
        email = datos.get('email', '').strip().lower()
        foto = datos.get('foto', '')

        if not email or not foto:
            return jsonify(ok=False, error='Datos incompletos'), 400

        encoding = obtener_encoding_facial_base64(foto)
        if not encoding:
            return jsonify(ok=False, error='No se detecto rostro'), 400

        client = supabase_admin if supabase_admin else supabase
        client.table('rostros_usuarios').upsert({
            'email': email,
            'encoding': encoding
        }, on_conflict='email').execute()

        try:
            subir_foto_rostro(email, foto)
        except Exception as e:
            print(f"Error subiendo JPG rostro: {e}")

        return jsonify(ok=True, mensaje='Rostro registrado')
    except Exception as e:
        print(f"Error registrando rostro admin: {e}")
        return jsonify(ok=False, error=str(e)), 500


# Alias para compatibilidad con el template actual
@app.route("/login", methods=["GET", "POST"])
def login():
    return redirect(url_for("auth_login"))


@app.route("/logout")
def logout():
    return redirect(url_for("auth_logout"))


# ==================== OCR ====================


def extraer_datos_imagen(ruta_imagen):
    img = Image.open(ruta_imagen)
    texto = pytesseract.image_to_string(img, lang="spa")

    datos = {"fecha": "", "valor": 0, "medio": "", "referencia": "", "observacion": ""}

    es_nequi = (
        "nequi" in texto.lower()
        or "envío realizado" in texto.lower()
        or "envio recibido" in texto.lower()
        or "envío recibido" in texto.lower()
    )
    es_banco_bogota = (
        "banco de bogotá" in texto.lower() or "banco de bogota" in texto.lower()
    )
    es_transfiya = "transfiya" in texto.lower()

    patron_valor = re.search(r"\$\s*([\d.,]+)", texto)
    if patron_valor:
        val_str = patron_valor.group(1).replace(".", "").replace(",", "")
        if len(val_str) > 2 and val_str.endswith("00"):
            val_str = val_str[:-2]
        try:
            datos["valor"] = int(val_str)
        except ValueError:
            pass

    patron_fecha_nequi = re.search(
        r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})", texto, re.IGNORECASE
    )
    patron_fecha_banco = re.search(
        r"(\w+)\s+(\d{1,2})\s+del\s+(\d{4})", texto, re.IGNORECASE
    )
    patron_fecha_alt = re.search(
        r"(\d{1,2})\s+de\s+(\w+)\s+del\s+(\d{4})", texto, re.IGNORECASE
    )

    if patron_fecha_nequi:
        dia = patron_fecha_nequi.group(1).zfill(2)
        mes = MESES.get(patron_fecha_nequi.group(2).lower(), "00")
        datos["fecha"] = f"{patron_fecha_nequi.group(3)}-{mes}-{dia}"
    elif patron_fecha_banco:
        mes = MESES.get(patron_fecha_banco.group(1).lower(), "00")
        dia = patron_fecha_banco.group(2).zfill(2)
        datos["fecha"] = f"{patron_fecha_banco.group(3)}-{mes}-{dia}"
    elif patron_fecha_alt:
        dia = patron_fecha_alt.group(1).zfill(2)
        mes = MESES.get(patron_fecha_alt.group(2).lower(), "00")
        datos["fecha"] = f"{patron_fecha_alt.group(3)}-{mes}-{dia}"

    if es_nequi and es_transfiya:
        datos["medio"] = "Banco de Bogota / Transfiya"
    elif es_nequi and es_banco_bogota:
        datos["medio"] = "Banco de Bogota / Nequi"
    elif es_nequi:
        datos["medio"] = "Nequi"
    elif es_transfiya:
        datos["medio"] = "Banco de Bogota / Transfiya"
    elif es_banco_bogota:
        datos["medio"] = "Banco de Bogota"
    else:
        datos["medio"] = "Otro"

    patron_ref_nequi = re.search(r"[Rr]eferencia\s*\n?\s*(\S+)", texto)
    patron_ref_banco = re.search(
        r"autorizaci[oó]n[:\s]*\n?\s*(\S+)", texto, re.IGNORECASE
    )
    patron_comprobante = re.search(r"[Cc]omprobante\s*\n?\s*(\d{10,})", texto)

    if patron_ref_nequi:
        ref = re.sub(r"^[^A-Za-z0-9]+", "", patron_ref_nequi.group(1).strip())
        datos["referencia"] = ref
    elif patron_ref_banco:
        datos["referencia"] = patron_ref_banco.group(1).strip()
    elif patron_comprobante:
        datos["referencia"] = patron_comprobante.group(1).strip()

    patron_para = re.search(r"(?:Para|De)\s*\n\s*(.+)", texto)
    patron_destino = re.search(r"[Cc]uenta destino[:\s]*\n?\s*(.+)", texto)
    patron_nota = re.search(r"[Nn]ota[:\s]*\n?\s*(.+)", texto)

    obs_partes = []
    if patron_para:
        obs_partes.append(patron_para.group(1).strip())
    if patron_destino:
        obs_partes.append("Destino: " + patron_destino.group(1).strip())
    if patron_nota:
        obs_partes.append("Nota: " + patron_nota.group(1).strip())
    datos["observacion"] = " - ".join(obs_partes) if obs_partes else ""

    return datos


# ==================== EXCEL ====================


def generar_excel(pagos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Financiero"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2E75B6", end_color="2E75B6", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    title_font = Font(bold=True, size=14)
    link_font = Font(color="0563C1", underline="single")
    rojo = Font(color="DC3545")
    verde = Font(color="28A745")

    ws.merge_cells("A1:J1")
    ws["A1"] = "Control Financiero"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    if pagos:
        fecha_desde = pagos[0].get("fecha", "")
        fecha_hasta = pagos[-1].get("fecha", "")
        ws["A2"] = f"Desde: {fecha_desde}  -  Hasta: {fecha_hasta}"
    else:
        ws["A2"] = ""
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "No.",
        "Fecha",
        "Concepto",
        "Debito",
        "Credito",
        "Saldo",
        "Medio de Pago",
        "Referencia",
        "Observaciones",
        "Comprobante",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    saldo = 0
    total_debitos = 0
    total_creditos = 0
    for i, pago in enumerate(pagos):
        row = i + 5
        tipo = pago.get("tipo", "egreso")
        valor = pago["valor"]
        debito = valor if tipo == "egreso" else 0
        credito = valor if tipo == "ingreso" else 0
        saldo += credito - debito
        total_debitos += debito
        total_creditos += credito

        ws.cell(row=row, column=1, value=i + 1).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=pago["fecha"]).border = border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=pago.get("concepto", "")).border = border

        cell_deb = ws.cell(row=row, column=4, value=debito if debito else None)
        cell_deb.border = border
        cell_deb.number_format = "$#,##0"
        cell_deb.alignment = Alignment(horizontal="right")
        if debito:
            cell_deb.font = rojo

        cell_cred = ws.cell(row=row, column=5, value=credito if credito else None)
        cell_cred.border = border
        cell_cred.number_format = "$#,##0"
        cell_cred.alignment = Alignment(horizontal="right")
        if credito:
            cell_cred.font = verde

        cell_saldo = ws.cell(row=row, column=6, value=saldo)
        cell_saldo.border = border
        cell_saldo.number_format = "$#,##0"
        cell_saldo.alignment = Alignment(horizontal="right")
        cell_saldo.font = verde if saldo >= 0 else rojo

        ws.cell(row=row, column=7, value=pago.get("medio", "")).border = border
        ws.cell(row=row, column=8, value=pago.get("referencia", "")).border = border
        ws.cell(row=row, column=9, value=pago.get("observacion", "")).border = border

        if pago.get("imagen"):
            cell_link = ws.cell(row=row, column=10, value="Ver comprobante")
            url_img = obtener_url_publica("comprobantes", pago["imagen"])
            cell_link.hyperlink = url_img
            cell_link.font = link_font
            cell_link.alignment = Alignment(horizontal="center")
            cell_link.border = border

    total_row = 5 + len(pagos)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).border = border

    cell_td = ws.cell(row=total_row, column=4, value=total_debitos)
    cell_td.number_format = "$#,##0"
    cell_td.font = Font(bold=True, size=11, color="DC3545")
    cell_td.alignment = Alignment(horizontal="right")

    cell_tc = ws.cell(row=total_row, column=5, value=total_creditos)
    cell_tc.number_format = "$#,##0"
    cell_tc.font = Font(bold=True, size=11, color="28A745")
    cell_tc.alignment = Alignment(horizontal="right")

    cell_ts = ws.cell(row=total_row, column=6, value=saldo)
    cell_ts.number_format = "$#,##0"
    cell_ts.font = Font(bold=True, size=11)
    cell_ts.alignment = Alignment(horizontal="right")

    ws.cell(row=total_row, column=9, value=f"{len(pagos)} movimientos")

    anchos = {
        "A": 8,
        "B": 14,
        "C": 22,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 28,
        "H": 28,
        "I": 40,
        "J": 18,
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, saldo


# ==================== CLASE MOVIMIENTO ====================


class Movimiento:
    def __init__(self, datos):
        self.id = datos.get("id", 0)
        self.fecha = datos.get("fecha", "")
        self.valor = datos.get("valor", 0)
        self.tipo = datos.get("tipo", "egreso")
        self.concepto = datos.get("concepto", "")
        self.medio = datos.get("medio", "")
        self.referencia = datos.get("referencia", "")
        self.observacion = datos.get("observacion", "")
        self.imagen = datos.get("imagen", "")
        self.imagen_url = (
            obtener_url_publica("comprobantes", self.imagen) if self.imagen else ""
        )
        self.latitud = datos.get("latitud")
        self.longitud = datos.get("longitud")

    @property
    def debito(self):
        return self.valor if self.tipo == "egreso" else 0

    @property
    def credito(self):
        return self.valor if self.tipo == "ingreso" else 0


def get_fechas(pagos_raw):
    if pagos_raw:
        return pagos_raw[0].get("fecha", ""), pagos_raw[-1].get("fecha", "")
    return "", ""


# ==================== RUTAS ====================


@app.route("/")
@login_requerido
def inicio():
    datos_raw = cargar_pagos()
    movimientos = [Movimiento(p) for p in datos_raw]
    total_ingresos = sum(m.credito for m in movimientos)
    total_egresos = sum(m.debito for m in movimientos)
    saldo = total_ingresos - total_egresos
    es_admin_flag = es_admin()
    usuario_id = session.get("usuario")
    vistos_resp = (
        supabase.table("pagos_vistos")
        .select("pago_id")
        .eq("usuario", usuario_id)
        .execute()
    )
    ids_vistos = {v["pago_id"] for v in vistos_resp.data}
    no_vistos = sum(1 for m in movimientos if m.imagen and m.id not in ids_vistos)

    pagos_json = [
        {
            "id": m.id,
            "fecha": m.fecha,
            "valor": m.valor,
            "tipo": m.tipo,
            "concepto": m.concepto,
            "medio": m.medio,
            "referencia": m.referencia,
            "observacion": m.observacion,
            "imagen": m.imagen,
            "imagen_url": m.imagen_url,
            "latitud": m.latitud,
            "longitud": m.longitud,
        }
        for m in movimientos
    ]

    return render_template(
        "paginas/inicio.html",
        movimientos=movimientos,
        total_ingresos=total_ingresos,
        total_egresos=total_egresos,
        saldo=saldo,
        es_admin=es_admin_flag,
        ids_vistos=ids_vistos,
        no_vistos=no_vistos,
        pagos_json=pagos_json,
        ids_vistos_list=list(ids_vistos),
        supabase_url=SUPABASE_URL,
        supabase_anon_key=SUPABASE_ANON_KEY,
    )


@app.route("/foto-usuario/<usuario>")
@login_requerido
def foto_usuario(usuario):
    try:
        url = obtener_url_publica("rostros", f"{usuario}.jpg")
        return redirect(url)
    except Exception:
        return "", 404


@app.route("/nuevo")
@login_requerido
def nuevo():
    return render_template("paginas/nuevo.html", datos=None)


@app.route("/chat", methods=["GET"])
@login_requerido
def chat():
    return render_template("paginas/chat.html")


@app.route("/graficas")
@login_requerido
def graficas():
    from collections import defaultdict

    pagos = cargar_pagos()

    # 1) Egresos por concepto
    por_concepto = defaultdict(int)
    for p in pagos:
        if p.get("tipo") == "egreso":
            clave = (p.get("concepto") or "Sin concepto").strip() or "Sin concepto"
            por_concepto[clave] += p.get("valor", 0)

    # 2) Egresos por medio
    por_medio = defaultdict(int)
    for p in pagos:
        if p.get("tipo") == "egreso":
            clave = (p.get("medio") or "Sin medio").strip() or "Sin medio"
            por_medio[clave] += p.get("valor", 0)

    # 3) Ingresos vs Egresos por mes
    por_mes = defaultdict(lambda: {"ingresos": 0, "egresos": 0})
    for p in pagos:
        fecha = p.get("fecha", "")
        mes = fecha[:7] if len(fecha) >= 7 else "sin-fecha"
        if p.get("tipo") == "ingreso":
            por_mes[mes]["ingresos"] += p.get("valor", 0)
        else:
            por_mes[mes]["egresos"] += p.get("valor", 0)
    meses_ord = sorted(por_mes.keys())

    # 4) Saldo acumulado
    saldos = []
    acum = 0
    for p in sorted(pagos, key=lambda x: (x.get("fecha", ""), x.get("id", 0))):
        valor = p.get("valor", 0)
        if p.get("tipo") == "ingreso":
            acum += valor
        else:
            acum -= valor
        saldos.append({"fecha": p.get("fecha", ""), "saldo": acum})

    # 5) Top 5 egresos mas grandes
    egresos_sorted = sorted(
        [p for p in pagos if p.get("tipo") == "egreso"],
        key=lambda x: x.get("valor", 0),
        reverse=True,
    )[:5]
    top_egresos = [
        {
            "etiqueta": f"{p.get('fecha','')} - {(p.get('concepto') or 'Sin concepto')[:30]}",
            "valor": p.get("valor", 0),
        }
        for p in egresos_sorted
    ]

    return render_template(
        "paginas/graficas.html",
        por_concepto=dict(por_concepto),
        por_medio=dict(por_medio),
        meses=meses_ord,
        mensual={m: dict(por_mes[m]) for m in meses_ord},
        saldos=saldos,
        top_egresos=top_egresos,
    )


@app.route("/api/chat", methods=["POST"])
@login_requerido
def api_chat():
    import requests as _req

    mensaje = (request.get_json() or {}).get("mensaje", "").strip()
    if not mensaje:
        return jsonify({"error": "Mensaje vacio"}), 400

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return jsonify({"error": "GEMINI_API_KEY no configurada"}), 500

    pagos = cargar_pagos()
    total_ingresos = sum(p["valor"] for p in pagos if p.get("tipo") == "ingreso")
    total_egresos = sum(p["valor"] for p in pagos if p.get("tipo") == "egreso")
    saldo = total_ingresos - total_egresos
    ultimos = pagos[-20:] if len(pagos) > 20 else pagos
    resumen = "\n".join(
        f"- {p.get('fecha')} | {p.get('tipo')} | ${p.get('valor'):,} | "
        f"{p.get('concepto','')} | {p.get('medio','')}"
        for p in ultimos
    )

    contexto = f"""Eres un asistente financiero del usuario {session.get('email', '')}.
Responde en espanol, breve y directo. Si preguntan sobre sus pagos, usa estos datos.
Si preguntan algo general (no de pagos), responde normal.

TOTALES:
- Ingresos: ${total_ingresos:,}
- Egresos: ${total_egresos:,}
- Saldo: ${saldo:,}
- Total movimientos: {len(pagos)}

ULTIMOS {len(ultimos)} MOVIMIENTOS:
{resumen}

Pregunta del usuario: {mensaje}
"""

    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent"
    try:
        r = _req.post(
            url,
            headers={"Content-Type": "application/json", "X-goog-api-key": api_key},
            json={"contents": [{"parts": [{"text": contexto}]}]},
            timeout=30,
        )
        data = r.json()
        if r.status_code != 200:
            return jsonify({"error": f"Gemini {r.status_code}: {data}"}), 500
        texto = data["candidates"][0]["content"]["parts"][0]["text"]
        return jsonify({"respuesta": texto})
    except Exception as e:
        return jsonify({"error": f"Error Gemini: {e}"}), 500


@app.route("/deploy", methods=["POST"])
def deploy():
    token = request.headers.get("X-Deploy-Token", "")
    expected = os.environ.get("DEPLOY_TOKEN", "")
    if not expected or token != expected:
        return jsonify({"error": "Token invalido"}), 401

    try:
        pull = subprocess.run(
            ["git", "-C", os.path.dirname(os.path.abspath(__file__)),
             "pull", "origin", "main"],
            capture_output=True, text=True, timeout=60,
        )
        wsgi = "/var/www/margoty_pythonanywhere_com_wsgi.py"
        if os.path.exists(wsgi):
            os.utime(wsgi, None)
        return jsonify({
            "ok": pull.returncode == 0,
            "stdout": pull.stdout,
            "stderr": pull.stderr,
        }), 200 if pull.returncode == 0 else 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/push-subscribe", methods=["POST"])
def push_subscribe():
    """Guardar suscripción de push notifications"""
    if "email" not in session:
        return jsonify({"error": "No autenticado"}), 401

    datos = request.get_json()
    endpoint = datos.get("endpoint")
    auth = datos.get("auth")
    p256dh = datos.get("p256dh")

    if not endpoint or not auth or not p256dh:
        return jsonify({"error": "Datos incompletos"}), 400

    usuario = session.get("email")
    if guardar_suscripcion_push(usuario, endpoint, auth, p256dh):
        return jsonify({"success": True}), 200
    else:
        return jsonify({"error": "Error guardando suscripción"}), 500


@app.route("/procesar-imagen", methods=["POST"])
@login_requerido
def procesar_imagen():
    if "imagen" not in request.files:
        flash("No se selecciono ninguna imagen.", "error")
        return redirect(url_for("nuevo"))

    archivo = request.files["imagen"]
    if archivo.filename == "":
        flash("No se selecciono ninguna imagen.", "error")
        return redirect(url_for("nuevo"))

    nombre = archivo.filename

    with tempfile.NamedTemporaryFile(
        suffix=os.path.splitext(nombre)[1], delete=False
    ) as tmp:
        ruta_temp = tmp.name
        archivo.save(ruta_temp)

    try:
        datos = extraer_datos_imagen(ruta_temp)
        datos["imagen"] = nombre

        with open(ruta_temp, "rb") as f:
            subir_imagen_storage("comprobantes", nombre, f.read())
    except Exception as e:
        flash(f"Error al leer la imagen con OCR: {e}", "error")
        return redirect(url_for("nuevo"))
    finally:
        os.remove(ruta_temp)

    try:
        pagos = cargar_pagos()
    except Exception as e:
        flash(f"Error cargando pagos: {e}", "error")
        return redirect(url_for("nuevo"))
    for p in pagos:
        if p.get("imagen") == nombre:
            flash(
                f"Esta imagen ya fue registrada (Fecha: {p['fecha']}, Valor: ${p['valor']:,.0f}).",
                "error",
            )
            return redirect(url_for("nuevo"))
        if datos["referencia"] and p.get("referencia") == datos["referencia"]:
            flash(
                f"Ya existe un pago con la misma referencia: {datos['referencia']}.",
                "error",
            )
            return redirect(url_for("nuevo"))

    return render_template("paginas/nuevo.html", datos=datos)


@app.route("/agregar", methods=["POST"])
@login_requerido
def agregar():
    fecha = request.form.get("fecha", "").strip()
    valor_str = (
        request.form.get("valor", "0")
        .strip()
        .replace(".", "")
        .replace(",", "")
        .replace("$", "")
    )
    tipo = request.form.get("tipo", "egreso").strip()
    concepto = request.form.get("concepto", "").strip()
    medio = request.form.get("medio", "").strip()
    referencia = request.form.get("referencia", "").strip()
    observacion = request.form.get("observacion", "").strip()
    imagen = request.form.get("imagen", "").strip()
    latitud_str = request.form.get("latitud", "").strip()
    longitud_str = request.form.get("longitud", "").strip()

    if not fecha or not valor_str:
        flash("Fecha y Valor son obligatorios.", "error")
        return redirect(url_for("nuevo"))

    try:
        valor = int(valor_str)
    except ValueError:
        flash("El valor debe ser un numero.", "error")
        return redirect(url_for("nuevo"))

    latitud = float(latitud_str) if latitud_str else None
    longitud = float(longitud_str) if longitud_str else None

    movimiento = {
        "fecha": fecha,
        "valor": valor,
        "tipo": tipo,
        "concepto": concepto,
        "medio": medio,
        "referencia": referencia,
        "observacion": observacion,
        "imagen": imagen,
        "latitud": latitud,
        "longitud": longitud,
    }

    guardar_pago(movimiento)

    pagos = cargar_pagos()
    etiqueta = "Ingreso" if tipo == "ingreso" else "Egreso"
    mensaje = f"{etiqueta} #{len(pagos)} por ${valor:,.0f} registrado correctamente."

    # Enviar notificación push a todos los usuarios
    try:
        enviar_notificacion_push(
            titulo="Nuevo Movimiento",
            cuerpo=f"{etiqueta} de ${valor:,.0f}: {concepto}"
        )
    except Exception as e:
        print(f"Error enviando notificación push: {e}")

    flash(mensaje, "exito")
    return redirect(url_for("inicio"))


@app.route("/eliminar/<int:pago_id>", methods=["POST"])
@login_requerido
def eliminar(pago_id):
    eliminar_pago(pago_id)
    flash("Movimiento eliminado.", "exito")
    return redirect(url_for("inicio"))


@app.route("/descargar-excel")
@login_requerido
def descargar_excel():
    pagos = cargar_pagos()
    if not pagos:
        flash("No hay movimientos registrados.", "error")
        return redirect(url_for("inicio"))
    buffer, saldo = generar_excel(pagos)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="control_financiero.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/correo")
@login_requerido
def correo():
    pagos = cargar_pagos()
    total_ingresos = sum(
        p["valor"] for p in pagos if p.get("tipo", "egreso") == "ingreso"
    )
    total_egresos = sum(
        p["valor"] for p in pagos if p.get("tipo", "egreso") == "egreso"
    )
    saldo = total_ingresos - total_egresos
    destinatarios = cargar_destinatarios()
    return render_template(
        "paginas/correo.html",
        total=saldo,
        num_pagos=len(pagos),
        remitente=CORREO_REMITENTE,
        destinatarios=destinatarios,
    )


@app.route("/agregar-destinatario", methods=["POST"])
@login_requerido
def agregar_destinatario():
    correo_nuevo = request.form.get("correo", "").strip().lower()
    if not correo_nuevo or "@" not in correo_nuevo:
        flash("Escribe un correo valido.", "error")
        return redirect(url_for("correo"))

    destinatarios = cargar_destinatarios()
    if correo_nuevo in destinatarios:
        flash("Ese correo ya esta en la lista.", "info")
        return redirect(url_for("correo"))

    agregar_destinatario_db(correo_nuevo)
    flash(f"Se agrego {correo_nuevo}.", "exito")
    return redirect(url_for("correo"))


@app.route("/quitar-destinatario", methods=["POST"])
@login_requerido
def quitar_destinatario():
    correo_quitar = request.form.get("correo", "")
    quitar_destinatario_db(correo_quitar)
    flash(f"Se quito {correo_quitar}.", "exito")
    return redirect(url_for("correo"))


@app.route("/enviar-correo", methods=["POST"])
@login_requerido
def enviar_correo():
    pagos = cargar_pagos()
    if not pagos:
        flash("No hay movimientos registrados.", "error")
        return redirect(url_for("correo"))

    destinatarios = cargar_destinatarios()
    if not destinatarios:
        flash("Agrega al menos un destinatario.", "error")
        return redirect(url_for("correo"))

    buffer_excel, saldo = generar_excel(pagos)

    try:
        msg = MIMEMultipart()
        msg["From"] = CORREO_REMITENTE
        msg["To"] = ", ".join(destinatarios)
        msg["Subject"] = (
            f"Control Financiero - {len(pagos)} movimientos - Saldo: ${saldo:,.0f}"
        )

        cuerpo = (
            f"Control Financiero (Supabase)\n"
            f"{'=' * 50}\n\n"
            f"Total movimientos: {len(pagos)}\n"
            f"Saldo actual: ${saldo:,.0f}\n"
            f"Fecha de envio: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
            f"Se adjunta un ZIP con el Excel y todos los comprobantes.\n"
        )
        msg.attach(MIMEText(cuerpo, "plain"))

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("control_financiero.xlsx", buffer_excel.read())
            for pago in pagos:
                if pago.get("imagen"):
                    try:
                        img_bytes = descargar_imagen_storage(
                            "comprobantes", pago["imagen"]
                        )
                        zf.writestr(pago["imagen"], img_bytes)
                    except Exception:
                        pass

        zip_buffer.seek(0)
        adjunto_zip = MIMEBase("application", "octet-stream")
        adjunto_zip.set_payload(zip_buffer.read())
        encoders.encode_base64(adjunto_zip)
        adjunto_zip.add_header(
            "Content-Disposition", "attachment", filename="control_financiero.zip"
        )
        msg.attach(adjunto_zip)

        servidor = smtplib.SMTP("smtp.gmail.com", 587)
        servidor.starttls()
        servidor.login(CORREO_REMITENTE, CORREO_CLAVE_APP)
        servidor.sendmail(CORREO_REMITENTE, destinatarios, msg.as_string())
        servidor.quit()

        destinos = ", ".join(destinatarios)
        flash(f"Correo enviado exitosamente a: {destinos}", "exito")

    except Exception as e:
        flash(f"Error al enviar correo: {e}", "error")

    return redirect(url_for("correo"))


@app.route("/marcar-visto/<int:pago_id>", methods=["POST"])
@login_requerido
def marcar_visto_ruta(pago_id):
    marcar_visto(session["usuario"], pago_id)
    no_vistos = obtener_no_vistos(session["usuario"])
    return jsonify(ok=True, no_vistos=no_vistos)


@app.route("/no-vistos")
@login_requerido
def no_vistos_ruta():
    no_vistos = obtener_no_vistos(session["usuario"])
    return jsonify(no_vistos=no_vistos)


@app.route("/sw.js")
def service_worker():
    return (
        app.send_static_file("sw.js"),
        200,
        {"Content-Type": "application/javascript"},
    )


if __name__ == "__main__":
    print("=" * 50)
    print("  Control Financiero (Supabase)")
    print("  Abre en tu navegador: http://localhost:5052")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5052, debug=True)
