import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pagos Liliana"

# Estilos
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
currency_format = '#,##0'
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
title_font = Font(bold=True, size=14)

# Titulo
ws.merge_cells('A1:H1')
ws['A1'] = 'Registro de Pagos - Liliana Patricia Moreno Amado'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:H2')
ws['A2'] = 'Nequi: 312 577 5938'
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['No. Cuota', 'Fecha', 'Valor Pago', 'Acumulado', 'Medio de Pago', 'Referencia', 'Observaciones', 'Comprobante']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Datos extraidos de las imagenes: [cuota, fecha, valor, acumulado, medio, referencia, observaciones, archivo_imagen]
pagos = [
    [1, '01/10/2025', 200000, 200000, 'Banco de Bogota / Transfiya', '02MW8pWCDKt5Cc6uC', '1er pago', '1er_pago_200000.jpeg'],
    [2, '16/10/2025', 200000, 400000, 'Banco de Bogota / Nequi', '9787877302341208712033130423392772', '2do pago - Liliana Patricia Moreno Amado', '2do_pago_200000_400000.jpeg'],
    [3, '01/11/2025', 200000, 600000, 'Nequi', 'M12114517', '3ra cuota - Enviado por Livio Martin Argoty Lucero', '3er_pago_200000_600000.jpeg'],
    [4, '15/11/2025', 200000, 800000, 'Nequi', 'M02839783', '4ta cuota', '4to_pago_200000_800000.jpeg'],
    [5, '30/11/2025', 200000, 1000000, 'Nequi', 'M18224863', '5ta cuota', '5to_pago_200000_1000000.jpeg'],
    [6, '17/12/2025', 200000, 1200000, 'Nequi', 'M06364242', '6ta cuota', '6to_pago_200000_1200000.jpeg'],
    [7, '01/01/2026', 200000, 1400000, 'Nequi', 'M07777160', '7ma cuota', '7ma_pago_200000_1400000.jpeg'],
    [8, '06/01/2026', 300000, 1700000, 'Nequi', 'M05642263', '8va cuota (7ma cuota en conversacion Nequi)', '8ta_pago_300000_1700000.jpeg'],
    [9, '31/01/2026', 200000, 1900000, 'Nequi', 'M05153763', '9na cuota', '9na_cuota_200000_1900000.jpeg'],
    [10, '16/02/2026', 200000, 2100000, 'Nequi', 'M22820548', '10ma cuota', '10_cuota_200000_2100000.jpeg'],
    [11, '28/02/2026', 200000, 2300000, 'Nequi', 'M12993665', '11va cuota', '11_pago_200000_2300000.jpeg'],
    [12, '14/03/2026', 200000, 2500000, 'Nequi', 'M15975635', '12va cuota', '12_pago_200000_2500000.jpeg'],
    [13, '31/03/2026', 120000, 2620000, 'Banco de Bogota', '000316', 'Transferencia a Ever Bco Popular - Nota: palmeras', '13_pago_200000_2500000.jpeg'],
]

link_font = Font(color="0563C1", underline="single")

for row_idx, pago in enumerate(pagos, 5):
    for col_idx, val in enumerate(pago[:7], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        if col_idx in (3, 4):  # Valor y Acumulado
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='right')
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal='center')
        elif col_idx == 2:
            cell.alignment = Alignment(horizontal='center')
    # Columna H: hipervinculo a la imagen (ruta relativa para portabilidad)
    img_file = pago[7]
    cell_link = ws.cell(row=row_idx, column=8, value='Ver comprobante')
    cell_link.hyperlink = './' + img_file
    cell_link.font = link_font
    cell_link.alignment = Alignment(horizontal='center')
    cell_link.border = border

# Fila de totales
total_row = 5 + len(pagos)
ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=11)
ws.cell(row=total_row, column=1).border = border
ws.cell(row=total_row, column=2).border = border
total_cell = ws.cell(row=total_row, column=3, value=sum(p[2] for p in pagos))
total_cell.number_format = '$#,##0'
total_cell.font = Font(bold=True, size=11)
total_cell.border = border
total_cell.alignment = Alignment(horizontal='right')
ws.cell(row=total_row, column=4).border = border
ws.cell(row=total_row, column=5).border = border
ws.cell(row=total_row, column=6).border = border
ws.cell(row=total_row, column=7, value='13 pagos realizados').border = border
ws.cell(row=total_row, column=8).border = border

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 50
ws.column_dimensions['H'].width = 20

wb.save('/home/margoty/Documentos/pagos_liliana/pagos_liliana.xlsx')
print("Excel generado: pagos_liliana.xlsx")
