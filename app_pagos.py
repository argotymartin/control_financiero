#!/usr/bin/env python3
"""
App de Registro de Pagos - Liliana Patricia Moreno Amado
Lee los datos automaticamente de las imagenes de comprobantes (OCR)
Funciona en Linux y Windows con Python 3 + openpyxl + pytesseract
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import zipfile
import pytesseract
from PIL import Image, ImageTk
import os
import platform
import shutil
import json
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Configurar Tesseract segun el sistema operativo
if platform.system() == 'Windows':
    rutas_tesseract = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        os.path.expanduser(r'~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'),
    ]
    for ruta in rutas_tesseract:
        if os.path.exists(ruta):
            pytesseract.pytesseract.tesseract_cmd = ruta
            break

RUTA_APP = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_EXCEL = os.path.join(RUTA_APP, 'pagos_liliana.xlsx')
ARCHIVO_DATOS = os.path.join(RUTA_APP, 'pagos_datos.json')

# Configuracion de correo
CORREO_REMITENTE = 'argoty.martin@gmail.com'
CORREO_CLAVE_APP = 'vbzs yrtk auhr quya'
ARCHIVO_DESTINATARIOS = os.path.join(RUTA_APP, 'destinatarios.json')


def cargar_destinatarios():
    if os.path.exists(ARCHIVO_DESTINATARIOS):
        with open(ARCHIVO_DESTINATARIOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    return ['martin.argoty@hotmail.com', 'lipamoan@hotmail.com']


def guardar_destinatarios(lista):
    with open(ARCHIVO_DESTINATARIOS, 'w', encoding='utf-8') as f:
        json.dump(lista, f, ensure_ascii=False, indent=2)

MESES = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
    'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
    'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
}

THUMB_W = 90
THUMB_H = 70


def extraer_datos_imagen(ruta_imagen):
    img = Image.open(ruta_imagen)
    texto = pytesseract.image_to_string(img, lang='spa')
    lineas = texto.strip().split('\n')
    texto_limpio = ' '.join(lineas)

    datos = {'fecha': '', 'valor': 0, 'medio': '', 'referencia': '', 'observacion': ''}

    es_nequi = 'nequi' in texto.lower() or 'envío realizado' in texto.lower() or 'envio recibido' in texto.lower() or 'envío recibido' in texto.lower()
    es_banco_bogota = 'banco de bogotá' in texto.lower() or 'banco de bogota' in texto.lower()
    es_transfiya = 'transfiya' in texto.lower()

    patron_valor = re.search(r'\$\s*([\d.,]+)', texto)
    if patron_valor:
        val_str = patron_valor.group(1).replace('.', '').replace(',', '')
        if len(val_str) > 2 and val_str.endswith('00'):
            val_str = val_str[:-2]
        try:
            datos['valor'] = int(val_str)
        except ValueError:
            pass

    patron_fecha_nequi = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})', texto, re.IGNORECASE)
    patron_fecha_banco = re.search(r'(\w+)\s+(\d{1,2})\s+del\s+(\d{4})', texto, re.IGNORECASE)
    patron_fecha_alt = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+del\s+(\d{4})', texto, re.IGNORECASE)

    if patron_fecha_nequi:
        dia = patron_fecha_nequi.group(1).zfill(2)
        mes = MESES.get(patron_fecha_nequi.group(2).lower(), '00')
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_nequi.group(3)}'
    elif patron_fecha_banco:
        mes = MESES.get(patron_fecha_banco.group(1).lower(), '00')
        dia = patron_fecha_banco.group(2).zfill(2)
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_banco.group(3)}'
    elif patron_fecha_alt:
        dia = patron_fecha_alt.group(1).zfill(2)
        mes = MESES.get(patron_fecha_alt.group(2).lower(), '00')
        datos['fecha'] = f'{dia}/{mes}/{patron_fecha_alt.group(3)}'

    if es_nequi and es_transfiya:
        datos['medio'] = 'Banco de Bogota / Transfiya'
    elif es_nequi and es_banco_bogota:
        datos['medio'] = 'Banco de Bogota / Nequi'
    elif es_nequi:
        datos['medio'] = 'Nequi'
    elif es_transfiya:
        datos['medio'] = 'Banco de Bogota / Transfiya'
    elif es_banco_bogota:
        datos['medio'] = 'Banco de Bogota'
    else:
        datos['medio'] = 'Otro'

    patron_ref_nequi = re.search(r'[Rr]eferencia\s*\n?\s*(\S+)', texto)
    patron_ref_banco = re.search(r'autorizaci[oó]n[:\s]*\n?\s*(\S+)', texto, re.IGNORECASE)
    patron_comprobante = re.search(r'[Cc]omprobante\s*\n?\s*(\d{10,})', texto)

    if patron_ref_nequi:
        ref = re.sub(r'^[^A-Za-z0-9]+', '', patron_ref_nequi.group(1).strip())
        datos['referencia'] = ref
    elif patron_ref_banco:
        datos['referencia'] = patron_ref_banco.group(1).strip()
    elif patron_comprobante:
        datos['referencia'] = patron_comprobante.group(1).strip()

    patron_para = re.search(r'(?:Para|De)\s*\n\s*(.+)', texto)
    patron_destino = re.search(r'[Cc]uenta destino[:\s]*\n?\s*(.+)', texto)
    patron_nota = re.search(r'[Nn]ota[:\s]*\n?\s*(.+)', texto)

    obs_partes = []
    if patron_para:
        obs_partes.append(patron_para.group(1).strip())
    if patron_destino:
        obs_partes.append('Destino: ' + patron_destino.group(1).strip())
    if patron_nota:
        obs_partes.append('Nota: ' + patron_nota.group(1).strip())
    datos['observacion'] = ' - '.join(obs_partes) if obs_partes else ''

    return datos, texto


def cargar_pagos():
    if os.path.exists(ARCHIVO_DATOS):
        with open(ARCHIVO_DATOS, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def guardar_pagos(pagos):
    with open(ARCHIVO_DATOS, 'w', encoding='utf-8') as f:
        json.dump(pagos, f, ensure_ascii=False, indent=2)


def generar_excel(pagos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos Liliana"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    link_font = Font(color="0563C1", underline="single")

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Registro de Pagos - Liliana Patricia Moreno Amado'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    if pagos:
        fecha_desde = pagos[0].get('fecha', '')
        fecha_hasta = pagos[-1].get('fecha', '')
        ws['A2'] = f'Desde: {fecha_desde}  -  Hasta: {fecha_hasta}'
    else:
        ws['A2'] = ''
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No. Cuota', 'Fecha', 'Valor Pago', 'Acumulado', 'Medio de Pago',
               'Referencia', 'Observaciones', 'Comprobante']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    acumulado = 0
    for i, pago in enumerate(pagos):
        row = i + 5
        valor = pago['valor']
        acumulado += valor
        datos = [i + 1, pago['fecha'], valor, acumulado, pago['medio'],
                 pago['referencia'], pago['observacion']]

        for col_idx, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
            if col_idx in (3, 4):
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in (1, 2):
                cell.alignment = Alignment(horizontal='center')

        if pago.get('imagen'):
            cell_link = ws.cell(row=row, column=8, value='Ver comprobante')
            cell_link.hyperlink = './' + pago['imagen']
            cell_link.font = link_font
            cell_link.alignment = Alignment(horizontal='center')
            cell_link.border = border

    total_row = 5 + len(pagos)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=11)
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).border = border
    total_cell = ws.cell(row=total_row, column=3, value=acumulado)
    total_cell.number_format = '$#,##0'
    total_cell.font = Font(bold=True, size=11)
    total_cell.alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=7, value=f'{len(pagos)} pagos realizados')

    anchos = {'A': 12, 'B': 14, 'C': 16, 'D': 16, 'E': 28, 'F': 38, 'G': 50, 'H': 20}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(ARCHIVO_EXCEL)
    return acumulado


def crear_miniatura(ruta_imagen):
    """Crea una miniatura que llena el espacio, recortando al centro si es necesario."""
    try:
        img = Image.open(ruta_imagen)
        # Escalar para que llene el area y recortar al centro
        img_w, img_h = img.size
        ratio_w = THUMB_W / img_w
        ratio_h = THUMB_H / img_h
        ratio = max(ratio_w, ratio_h)
        nuevo_w = int(img_w * ratio)
        nuevo_h = int(img_h * ratio)
        img = img.resize((nuevo_w, nuevo_h), Image.LANCZOS)
        # Recortar al centro
        left = (nuevo_w - THUMB_W) // 2
        top = (nuevo_h - THUMB_H) // 2
        img = img.crop((left, top, left + THUMB_W, top + THUMB_H))
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


def mostrar_imagen_completa(ruta_imagen):
    """Abre una ventana nueva con la imagen en tamano completo."""
    if not os.path.exists(ruta_imagen):
        messagebox.showerror("Error", f"No se encontro la imagen:\n{ruta_imagen}")
        return

    ventana = tk.Toplevel()
    ventana.title(os.path.basename(ruta_imagen))

    img = Image.open(ruta_imagen)

    # Escalar si es muy grande para la pantalla
    screen_w = ventana.winfo_screenwidth() - 100
    screen_h = ventana.winfo_screenheight() - 100
    img_w, img_h = img.size

    if img_w > screen_w or img_h > screen_h:
        ratio = min(screen_w / img_w, screen_h / img_h)
        nuevo_w = int(img_w * ratio)
        nuevo_h = int(img_h * ratio)
        img = img.resize((nuevo_w, nuevo_h), Image.LANCZOS)

    foto = ImageTk.PhotoImage(img)

    # Canvas con scroll para imagenes grandes
    canvas = tk.Canvas(ventana, width=img.size[0], height=img.size[1])
    canvas.pack(fill='both', expand=True)
    canvas.create_image(0, 0, anchor='nw', image=foto)
    canvas._foto = foto  # Mantener referencia

    ventana.geometry(f"{img.size[0]}x{img.size[1]}")


def verificar_consistencia(pagos):
    """Verifica inconsistencias entre el JSON y las imagenes en la carpeta."""
    problemas = []

    # Imagenes registradas en el JSON
    imagenes_json = set()
    for i, pago in enumerate(pagos):
        img = pago.get('imagen', '')
        if img:
            imagenes_json.add(img)
            ruta = os.path.join(RUTA_APP, img)
            if not os.path.exists(ruta):
                problemas.append(f"  FALTA ARCHIVO: Pago #{i+1} ({pago['fecha']} - ${pago['valor']:,.0f}) "
                                f"referencia imagen \"{img}\" pero no existe en la carpeta.")

    # Imagenes en la carpeta que no estan en el JSON
    extensiones = ('.jpeg', '.jpg', '.png', '.bmp')
    for archivo in sorted(os.listdir(RUTA_APP)):
        if archivo.lower().endswith(extensiones):
            if archivo not in imagenes_json:
                problemas.append(f"  SIN REGISTRAR: La imagen \"{archivo}\" esta en la carpeta "
                                f"pero no tiene pago asociado en el sistema.")

    # Pagos sin imagen
    for i, pago in enumerate(pagos):
        if not pago.get('imagen'):
            problemas.append(f"  SIN IMAGEN: Pago #{i+1} ({pago['fecha']} - ${pago['valor']:,.0f}) "
                            f"no tiene comprobante de imagen asociado.")

    return problemas


class AppPagos:
    def __init__(self, root):
        self.root = root
        self.root.title("Registro de Pagos - Liliana Moreno")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        self.root.minsize(1000, 700)

        self.pagos = cargar_pagos()
        self.miniaturas = {}  # Cache de miniaturas para evitar garbage collection

        self.crear_interfaz()
        self.actualizar_tabla()
        self.verificar_al_iniciar()

    def crear_interfaz(self):
        # Frame superior - cargar imagen
        frame_cargar = ttk.LabelFrame(self.root, text="Cargar Comprobante de Pago", padding=15)
        frame_cargar.pack(fill='x', padx=15, pady=8)

        fila_btn = ttk.Frame(frame_cargar)
        fila_btn.pack(fill='x', pady=5)

        self.btn_cargar = tk.Button(fila_btn, text="\U0001F4F7  Seleccionar Imagen de Comprobante...",
                                     command=self.cargar_imagen, bg='#0078D4', fg='white',
                                     font=('Arial', 11, 'bold'), padx=15, pady=8, cursor='hand2')
        self.btn_cargar.pack(side='left', padx=5)

        self.lbl_estado = ttk.Label(fila_btn, text="Selecciona una imagen y los datos se extraen automaticamente",
                                     foreground='gray', font=('Arial', 10))
        self.lbl_estado.pack(side='left', padx=15)

        # Frame de datos extraidos
        self.frame_datos = ttk.LabelFrame(self.root, text="Datos Extraidos (puedes corregir antes de agregar)",
                                           padding=15)
        self.frame_datos.pack(fill='x', padx=15, pady=5)

        campos = ttk.Frame(self.frame_datos)
        campos.pack(fill='x')

        # Configurar columnas para que se expandan
        for c in range(6):
            campos.columnconfigure(c, weight=1)

        fuente_label = ('Arial', 10, 'bold')
        fuente_entry = ('Arial', 11)

        ttk.Label(campos, text="Fecha:", font=fuente_label).grid(row=0, column=0, sticky='e', padx=8, pady=6)
        self.entry_fecha = ttk.Entry(campos, width=18, font=fuente_entry)
        self.entry_fecha.grid(row=0, column=1, padx=8, pady=6, sticky='ew')

        ttk.Label(campos, text="Valor $:", font=fuente_label).grid(row=0, column=2, sticky='e', padx=8, pady=6)
        self.entry_valor = ttk.Entry(campos, width=18, font=fuente_entry)
        self.entry_valor.grid(row=0, column=3, padx=8, pady=6, sticky='ew')

        ttk.Label(campos, text="Medio:", font=fuente_label).grid(row=0, column=4, sticky='e', padx=8, pady=6)
        self.entry_medio = ttk.Entry(campos, width=28, font=fuente_entry)
        self.entry_medio.grid(row=0, column=5, padx=8, pady=6, sticky='ew')

        ttk.Label(campos, text="Referencia:", font=fuente_label).grid(row=1, column=0, sticky='e', padx=8, pady=6)
        self.entry_ref = ttk.Entry(campos, width=30, font=fuente_entry)
        self.entry_ref.grid(row=1, column=1, columnspan=2, sticky='ew', padx=8, pady=6)

        ttk.Label(campos, text="Observacion:", font=fuente_label).grid(row=1, column=3, sticky='e', padx=8, pady=6)
        self.entry_obs = ttk.Entry(campos, width=45, font=fuente_entry)
        self.entry_obs.grid(row=1, column=4, columnspan=2, sticky='ew', padx=8, pady=6)

        fila_confirmar = ttk.Frame(self.frame_datos)
        fila_confirmar.pack(fill='x', pady=8)

        self.btn_agregar = tk.Button(fila_confirmar, text="\u2714 CONFIRMAR Y AGREGAR PAGO",
                                      command=self.confirmar_pago, state='disabled',
                                      bg='#28A745', fg='white', font=('Arial', 11, 'bold'),
                                      padx=15, pady=6, cursor='hand2',
                                      disabledforeground='#999999')
        self.btn_agregar.pack(side='right', padx=5)

        self.lbl_imagen_nombre = ttk.Label(fila_confirmar, text="", foreground='blue',
                                            font=('Arial', 10))
        self.lbl_imagen_nombre.pack(side='left', padx=5)

        # Frame tabla con miniaturas
        frame_tabla = ttk.LabelFrame(self.root,
                                      text="Historial de Pagos  (clic en miniatura = ver imagen completa  |  clic en fila = seleccionar)",
                                      padding=8)
        frame_tabla.pack(fill='both', expand=True, padx=15, pady=5)

        # Canvas con scroll para la tabla de pagos
        self.canvas_tabla = tk.Canvas(frame_tabla, bg='white')
        scrollbar_v = ttk.Scrollbar(frame_tabla, orient='vertical', command=self.canvas_tabla.yview)
        self.canvas_tabla.configure(yscrollcommand=scrollbar_v.set)

        scrollbar_v.pack(side='right', fill='y')
        self.canvas_tabla.pack(side='left', fill='both', expand=True)

        self.frame_interior = tk.Frame(self.canvas_tabla, bg='white')
        self.canvas_window = self.canvas_tabla.create_window((0, 0), window=self.frame_interior, anchor='nw')

        # Hacer que el frame interior ocupe todo el ancho del canvas
        def on_canvas_configure(event):
            self.canvas_tabla.itemconfig(self.canvas_window, width=event.width)
        self.canvas_tabla.bind('<Configure>', on_canvas_configure)

        self.frame_interior.bind('<Configure>',
                                  lambda e: self.canvas_tabla.configure(scrollregion=self.canvas_tabla.bbox('all')))

        # Scroll con rueda del mouse
        self.canvas_tabla.bind_all('<MouseWheel>',
                                    lambda e: self.canvas_tabla.yview_scroll(int(-1 * (e.delta / 120)), 'units'))
        self.canvas_tabla.bind_all('<Button-4>',
                                    lambda e: self.canvas_tabla.yview_scroll(-1, 'units'))
        self.canvas_tabla.bind_all('<Button-5>',
                                    lambda e: self.canvas_tabla.yview_scroll(1, 'units'))

        # Frame inferior
        frame_acciones = ttk.Frame(self.root, padding=8)
        frame_acciones.pack(fill='x', padx=15, pady=8)

        self.lbl_total = ttk.Label(frame_acciones, text="Total: $0", font=('Arial', 16, 'bold'))
        self.lbl_total.pack(side='left', padx=15)

        self.btn_correo = tk.Button(frame_acciones, text="\u2709  Enviar por Correo",
                                      command=self.enviar_correo, bg='#D44638', fg='white',
                                      font=('Arial', 12, 'bold'), padx=20, pady=8, cursor='hand2')
        self.btn_correo.pack(side='right', padx=8)

        self.btn_excel = tk.Button(frame_acciones, text="\U0001F4CA  Generar Excel",
                                    command=self.generar, bg='#217346', fg='white',
                                    font=('Arial', 12, 'bold'), padx=20, pady=8, cursor='hand2')
        self.btn_excel.pack(side='right', padx=8)

        self.btn_eliminar = tk.Button(frame_acciones, text="\u2716  Eliminar Seleccionado",
                                       command=self.eliminar_seleccionado, bg='#CC0000', fg='white',
                                       font=('Arial', 12, 'bold'), padx=20, pady=8, cursor='hand2')
        self.btn_eliminar.pack(side='right', padx=8)

        self.imagen_pendiente = None
        self.fila_seleccionada = None

    def cargar_imagen(self):
        archivo = filedialog.askopenfilename(
            title="Seleccionar comprobante de pago",
            filetypes=[("Imagenes", "*.jpeg *.jpg *.png *.bmp"), ("Todos", "*.*")]
        )
        if not archivo:
            return

        self.lbl_estado.config(text="Leyendo imagen con OCR...", foreground='orange')
        self.root.update()

        try:
            datos, texto_ocr = extraer_datos_imagen(archivo)
        except Exception as e:
            messagebox.showerror("Error OCR", f"No se pudo leer la imagen:\n{e}")
            self.lbl_estado.config(text="Error al leer imagen", foreground='red')
            return

        self.entry_fecha.delete(0, 'end')
        self.entry_fecha.insert(0, datos['fecha'])
        self.entry_valor.delete(0, 'end')
        self.entry_valor.insert(0, str(datos['valor']))
        self.entry_medio.delete(0, 'end')
        self.entry_medio.insert(0, datos['medio'])
        self.entry_ref.delete(0, 'end')
        self.entry_ref.insert(0, datos['referencia'])
        self.entry_obs.delete(0, 'end')
        self.entry_obs.insert(0, datos['observacion'])

        self.imagen_pendiente = archivo
        nombre = os.path.basename(archivo)

        # Validar duplicados
        duplicado = False
        for p in self.pagos:
            if p.get('imagen') == nombre:
                duplicado = True
                messagebox.showwarning("Pago duplicado",
                    f"Esta imagen ya fue registrada.\n\n"
                    f"Pago existente:\n"
                    f"  Fecha: {p['fecha']}\n"
                    f"  Valor: ${p['valor']:,.0f}\n"
                    f"  Referencia: {p['referencia']}")
                break
            if datos['referencia'] and p.get('referencia') == datos['referencia']:
                duplicado = True
                messagebox.showwarning("Pago duplicado",
                    f"Ya existe un pago con la misma referencia: {datos['referencia']}\n\n"
                    f"Pago existente:\n"
                    f"  Fecha: {p['fecha']}\n"
                    f"  Valor: ${p['valor']:,.0f}\n"
                    f"  Imagen: {p.get('imagen', '')}")
                break

        if duplicado:
            self.lbl_imagen_nombre.config(text=f"DUPLICADO: {nombre}")
            self.lbl_estado.config(text="Este pago ya fue registrado.", foreground='red')
            self.btn_agregar.config(state='disabled')
            return

        self.lbl_imagen_nombre.config(text=f"Imagen: {nombre}")
        self.lbl_estado.config(text="Datos extraidos. Revisa y confirma.", foreground='green')
        self.btn_agregar.config(state='normal')

    def confirmar_pago(self):
        fecha = self.entry_fecha.get().strip()
        valor_str = self.entry_valor.get().strip().replace('.', '').replace(',', '').replace('$', '')
        medio = self.entry_medio.get().strip()
        referencia = self.entry_ref.get().strip()
        observacion = self.entry_obs.get().strip()

        if not fecha or not valor_str:
            messagebox.showwarning("Campos requeridos", "Al menos Fecha y Valor son necesarios.")
            return

        try:
            valor = int(valor_str)
        except ValueError:
            messagebox.showerror("Error", "El valor debe ser un numero (ej: 200000)")
            return

        nombre_imagen = ''
        if self.imagen_pendiente:
            nombre_imagen = os.path.basename(self.imagen_pendiente)
            destino = os.path.join(RUTA_APP, nombre_imagen)
            origen = self.imagen_pendiente
            if os.path.abspath(origen) != os.path.abspath(destino):
                shutil.copy2(origen, destino)

        pago = {
            'fecha': fecha, 'valor': valor, 'medio': medio,
            'referencia': referencia, 'observacion': observacion,
            'imagen': nombre_imagen
        }

        self.pagos.append(pago)
        guardar_pagos(self.pagos)
        self.actualizar_tabla()
        self.limpiar_formulario()
        messagebox.showinfo("Pago agregado", f"Pago #{len(self.pagos)} por ${valor:,.0f} registrado.")

    def _color_fila(self, row_num, color):
        """Cambia el color de fondo de toda una fila."""
        for widget in self.frame_interior.grid_slaves(row=row_num):
            if isinstance(widget, tk.Label):
                widget.configure(bg=color)
            elif isinstance(widget, tk.Frame):
                widget.configure(bg=color)
                for child in widget.winfo_children():
                    child.configure(bg=color)

    def _get_bg_fila(self, indice):
        """Retorna el color de fondo original de una fila."""
        if self.fila_seleccionada == indice:
            return '#CCE5FF'
        return 'white' if indice % 2 == 0 else '#F0F0F0'

    def hover_enter(self, indice):
        """Resalta la fila al pasar el mouse."""
        if self.fila_seleccionada != indice:
            self._color_fila(indice + 1, '#E8F4FD')

    def hover_leave(self, indice):
        """Restaura el color al salir el mouse."""
        if self.fila_seleccionada != indice:
            self._color_fila(indice + 1, self._get_bg_fila(indice))

    def seleccionar_fila(self, indice):
        """Marca una fila como seleccionada."""
        # Deseleccionar anterior
        if self.fila_seleccionada is not None:
            prev = self.fila_seleccionada
            self.fila_seleccionada = None  # Limpiar antes para que _get_bg_fila funcione bien
            self._color_fila(prev + 1, self._get_bg_fila(prev))

        self.fila_seleccionada = indice
        self._color_fila(indice + 1, '#CCE5FF')

    def eliminar_seleccionado(self):
        if self.fila_seleccionada is None:
            messagebox.showinfo("Info", "Selecciona un pago de la tabla para eliminar.")
            return

        indice = self.fila_seleccionada
        if indice < 0 or indice >= len(self.pagos):
            return

        pago = self.pagos[indice]
        confirmar = messagebox.askyesno("Confirmar eliminacion",
                                         f"Eliminar pago #{indice + 1}?\n\n"
                                         f"  Fecha: {pago['fecha']}\n"
                                         f"  Valor: ${pago['valor']:,.0f}\n"
                                         f"  Medio: {pago['medio']}\n"
                                         f"  Referencia: {pago['referencia']}\n"
                                         f"  Imagen: {pago.get('imagen', '')}")
        if confirmar:
            self.pagos.pop(indice)
            self.fila_seleccionada = None
            guardar_pagos(self.pagos)
            self.actualizar_tabla()

    def actualizar_tabla(self):
        # Limpiar tabla
        for widget in self.frame_interior.winfo_children():
            widget.destroy()
        self.miniaturas.clear()
        self.fila_seleccionada = None

        # Headers
        headers = ['#', 'Imagen', 'Fecha', 'Valor', 'Acumulado', 'Medio', 'Referencia']
        pesos =   [1,   2,        3,       3,       3,           5,       5]
        header_bg = '#2E75B6'

        for col, (header, peso) in enumerate(zip(headers, pesos)):
            self.frame_interior.columnconfigure(col, weight=peso)
            lbl = tk.Label(self.frame_interior, text=header, font=('Arial', 11, 'bold'),
                          bg=header_bg, fg='white', anchor='center',
                          padx=8, pady=8, relief='ridge')
            lbl.grid(row=0, column=col, sticky='nsew')

        # Filas de datos
        acumulado = 0
        for i, pago in enumerate(self.pagos):
            acumulado += pago['valor']
            row = i + 1
            bg = 'white' if i % 2 == 0 else '#F0F0F0'

            datos_fila = [
                str(i + 1),
                None,  # Placeholder para miniatura
                pago['fecha'],
                f"${pago['valor']:,.0f}",
                f"${acumulado:,.0f}",
                pago['medio'],
                pago['referencia']
            ]

            for col, dato in enumerate(datos_fila):
                if col == 1:  # Columna de miniatura
                    frame_img = tk.Frame(self.frame_interior, bg=bg, relief='ridge',
                                          borderwidth=1, padx=2, pady=2)
                    frame_img.grid(row=row, column=col, sticky='nsew')

                    ruta_img = os.path.join(RUTA_APP, pago.get('imagen', ''))
                    if pago.get('imagen') and os.path.exists(ruta_img):
                        thumb = crear_miniatura(ruta_img)
                        if thumb:
                            self.miniaturas[i] = thumb
                            lbl_img = tk.Label(frame_img, image=thumb, bg=bg, cursor='hand2')
                            lbl_img.pack(fill='both', expand=True)
                            lbl_img.bind('<Button-1>',
                                          lambda e, r=ruta_img: mostrar_imagen_completa(r))
                            # Hover en miniatura
                            lbl_img.bind('<Enter>', lambda e, idx=i: self.hover_enter(idx))
                            lbl_img.bind('<Leave>', lambda e, idx=i: self.hover_leave(idx))
                    else:
                        lbl_sin = tk.Label(frame_img, text="--", bg=bg, fg='gray')
                        lbl_sin.pack(expand=True)
                        lbl_sin.bind('<Enter>', lambda e, idx=i: self.hover_enter(idx))
                        lbl_sin.bind('<Leave>', lambda e, idx=i: self.hover_leave(idx))

                    frame_img.bind('<Button-1>', lambda e, idx=i: self.seleccionar_fila(idx))
                    frame_img.bind('<Enter>', lambda e, idx=i: self.hover_enter(idx))
                    frame_img.bind('<Leave>', lambda e, idx=i: self.hover_leave(idx))
                else:
                    anchor = 'e' if col in (3, 4) else 'center'
                    lbl = tk.Label(self.frame_interior, text=dato, font=('Arial', 10),
                                  bg=bg, anchor=anchor, padx=8, pady=10, relief='ridge')
                    lbl.grid(row=row, column=col, sticky='nsew')
                    lbl.bind('<Button-1>', lambda e, idx=i: self.seleccionar_fila(idx))
                    lbl.bind('<Enter>', lambda e, idx=i: self.hover_enter(idx))
                    lbl.bind('<Leave>', lambda e, idx=i: self.hover_leave(idx))

        self.lbl_total.config(text=f"Total: ${acumulado:,.0f}  ({len(self.pagos)} pagos)")

    def limpiar_formulario(self):
        self.entry_fecha.delete(0, 'end')
        self.entry_valor.delete(0, 'end')
        self.entry_medio.delete(0, 'end')
        self.entry_ref.delete(0, 'end')
        self.entry_obs.delete(0, 'end')
        self.imagen_pendiente = None
        self.lbl_imagen_nombre.config(text="")
        self.lbl_estado.config(text="Selecciona una imagen y los datos se extraen automaticamente",
                               foreground='gray')
        self.btn_agregar.config(state='disabled')

    def verificar_al_iniciar(self):
        """Muestra inconsistencias al iniciar la app."""
        problemas = verificar_consistencia(self.pagos)
        if problemas:
            titulo = f"Se encontraron {len(problemas)} inconsistencia(s)"
            mensaje = "REPORTE DE CONSISTENCIA\n"
            mensaje += "=" * 40 + "\n\n"
            mensaje += "\n\n".join(problemas)
            mensaje += "\n\n" + "=" * 40
            mensaje += "\nRevisa estos problemas antes de continuar."

            # Ventana personalizada para mostrar el reporte
            ventana = tk.Toplevel(self.root)
            ventana.title(titulo)
            ventana.geometry("700x400")
            ventana.configure(bg='#FFF3CD')

            lbl_titulo = tk.Label(ventana, text=f"\u26A0  {titulo}",
                                   font=('Arial', 14, 'bold'), bg='#FFF3CD', fg='#856404')
            lbl_titulo.pack(pady=10)

            frame_texto = tk.Frame(ventana)
            frame_texto.pack(fill='both', expand=True, padx=15, pady=5)

            texto = tk.Text(frame_texto, wrap='word', font=('Arial', 10),
                           bg='white', fg='#333333', padx=10, pady=10)
            scroll = ttk.Scrollbar(frame_texto, command=texto.yview)
            texto.configure(yscrollcommand=scroll.set)

            scroll.pack(side='right', fill='y')
            texto.pack(side='left', fill='both', expand=True)

            for prob in problemas:
                if 'FALTA ARCHIVO' in prob:
                    texto.insert('end', prob + '\n\n', 'rojo')
                elif 'SIN REGISTRAR' in prob:
                    texto.insert('end', prob + '\n\n', 'naranja')
                elif 'SIN IMAGEN' in prob:
                    texto.insert('end', prob + '\n\n', 'gris')

            texto.tag_config('rojo', foreground='#CC0000')
            texto.tag_config('naranja', foreground='#E67E00')
            texto.tag_config('gris', foreground='#666666')
            texto.config(state='disabled')

            tk.Button(ventana, text="Entendido", command=ventana.destroy,
                     bg='#856404', fg='white', font=('Arial', 11, 'bold'),
                     padx=20, pady=5, cursor='hand2').pack(pady=10)

            ventana.transient(self.root)
            ventana.grab_set()

    def enviar_correo(self):
        if not self.pagos:
            messagebox.showwarning("Sin datos", "No hay pagos registrados.")
            return

        # Generar el Excel primero
        acumulado = generar_excel(self.pagos)

        # Abrir ventana de destinatarios
        ventana = tk.Toplevel(self.root)
        ventana.title("Enviar Excel por Correo")
        ventana.geometry("550x480")
        ventana.configure(bg='#F5F5F5')
        ventana.resizable(False, False)
        ventana.transient(self.root)
        ventana.grab_set()

        # Info del envio
        frame_info = tk.Frame(ventana, bg='#E8F4FD', padx=15, pady=10)
        frame_info.pack(fill='x', padx=15, pady=(15, 5))

        tk.Label(frame_info, text="Archivo: pagos_liliana.xlsx",
                 font=('Arial', 10, 'bold'), bg='#E8F4FD').pack(anchor='w')
        tk.Label(frame_info, text=f"Total: ${acumulado:,.0f}  -  {len(self.pagos)} pagos",
                 font=('Arial', 10), bg='#E8F4FD').pack(anchor='w')
        tk.Label(frame_info, text=f"Remitente: {CORREO_REMITENTE}",
                 font=('Arial', 10), bg='#E8F4FD').pack(anchor='w')

        # Lista de destinatarios
        tk.Label(ventana, text="Destinatarios:", font=('Arial', 11, 'bold'),
                 bg='#F5F5F5').pack(anchor='w', padx=15, pady=(10, 5))

        frame_lista = tk.Frame(ventana, bg='#F5F5F5')
        frame_lista.pack(fill='both', expand=True, padx=15)

        listbox = tk.Listbox(frame_lista, font=('Arial', 11), height=8,
                             selectmode='single', activestyle='none')
        scroll = ttk.Scrollbar(frame_lista, command=listbox.yview)
        listbox.configure(yscrollcommand=scroll.set)
        scroll.pack(side='right', fill='y')
        listbox.pack(side='left', fill='both', expand=True)

        destinatarios = cargar_destinatarios()
        for d in destinatarios:
            listbox.insert('end', d)

        # Frame para agregar correo
        frame_agregar = tk.Frame(ventana, bg='#F5F5F5')
        frame_agregar.pack(fill='x', padx=15, pady=8)

        entry_nuevo = ttk.Entry(frame_agregar, font=('Arial', 11))
        entry_nuevo.pack(side='left', fill='x', expand=True, padx=(0, 8))
        entry_nuevo.insert(0, 'correo@ejemplo.com')
        entry_nuevo.bind('<FocusIn>', lambda e: entry_nuevo.delete(0, 'end')
                         if entry_nuevo.get() == 'correo@ejemplo.com' else None)

        def agregar():
            correo = entry_nuevo.get().strip().lower()
            if not correo or correo == 'correo@ejemplo.com':
                return
            if '@' not in correo or '.' not in correo:
                messagebox.showwarning("Correo invalido",
                    "Escribe un correo valido (ejemplo: nombre@correo.com)", parent=ventana)
                return
            if correo in listbox.get(0, 'end'):
                messagebox.showinfo("Duplicado", "Ese correo ya esta en la lista.", parent=ventana)
                return
            listbox.insert('end', correo)
            entry_nuevo.delete(0, 'end')

        def quitar():
            sel = listbox.curselection()
            if not sel:
                messagebox.showinfo("Info", "Selecciona un correo de la lista para quitar.",
                                    parent=ventana)
                return
            listbox.delete(sel[0])

        tk.Button(frame_agregar, text="+ Agregar", command=agregar,
                  bg='#0078D4', fg='white', font=('Arial', 10, 'bold'),
                  padx=10, cursor='hand2').pack(side='left', padx=(0, 4))

        tk.Button(frame_agregar, text="- Quitar", command=quitar,
                  bg='#CC0000', fg='white', font=('Arial', 10, 'bold'),
                  padx=10, cursor='hand2').pack(side='left')

        # Botones de accion
        frame_botones = tk.Frame(ventana, bg='#F5F5F5')
        frame_botones.pack(fill='x', padx=15, pady=(5, 15))

        def cancelar():
            ventana.destroy()

        def enviar():
            lista = list(listbox.get(0, 'end'))
            if not lista:
                messagebox.showwarning("Sin destinatarios",
                    "Agrega al menos un destinatario.", parent=ventana)
                return

            guardar_destinatarios(lista)
            ventana.destroy()

            self.lbl_estado.config(text="Enviando correo...", foreground='orange')
            self.root.update()

            try:
                msg = MIMEMultipart()
                msg['From'] = CORREO_REMITENTE
                msg['To'] = ', '.join(lista)
                msg['Subject'] = (f'Registro de Pagos Liliana - '
                                  f'{len(self.pagos)} pagos - ${acumulado:,.0f}')

                cuerpo = (
                    f"Registro de Pagos - Liliana Patricia Moreno Amado\n"
                    f"{'=' * 50}\n\n"
                    f"Total pagos: {len(self.pagos)}\n"
                    f"Total acumulado: ${acumulado:,.0f}\n"
                    f"Fecha de envio: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
                    f"Se adjunta un ZIP con el Excel y todos los comprobantes.\n"
                    f"Descomprime el ZIP en una carpeta y abre el Excel desde ahi.\n"
                    f"Los enlaces 'Ver comprobante' abren las imagenes directamente.\n"
                )
                msg.attach(MIMEText(cuerpo, 'plain'))

                # Crear ZIP con Excel + imagenes
                archivo_zip = os.path.join(RUTA_APP, 'pagos_liliana.zip')
                with zipfile.ZipFile(archivo_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(ARCHIVO_EXCEL, 'pagos_liliana.xlsx')
                    for pago in self.pagos:
                        if pago.get('imagen'):
                            ruta_img = os.path.join(RUTA_APP, pago['imagen'])
                            if os.path.exists(ruta_img):
                                zf.write(ruta_img, pago['imagen'])

                with open(archivo_zip, 'rb') as f:
                    adjunto_zip = MIMEBase('application', 'octet-stream')
                    adjunto_zip.set_payload(f.read())
                    encoders.encode_base64(adjunto_zip)
                    adjunto_zip.add_header('Content-Disposition',
                                           'attachment', filename='pagos_liliana.zip')
                    msg.attach(adjunto_zip)

                servidor = smtplib.SMTP('smtp.gmail.com', 587)
                servidor.starttls()
                servidor.login(CORREO_REMITENTE, CORREO_CLAVE_APP)
                servidor.sendmail(CORREO_REMITENTE, lista, msg.as_string())
                servidor.quit()

                self.lbl_estado.config(text="Correo enviado exitosamente!",
                                       foreground='green')
                destinos = '\n'.join(f'  - {d}' for d in lista)
                messagebox.showinfo("Correo enviado",
                    f"El archivo Excel fue enviado a:\n\n{destinos}")

            except Exception as e:
                self.lbl_estado.config(text="Error al enviar correo", foreground='red')
                messagebox.showerror("Error de correo",
                    f"No se pudo enviar el correo:\n\n{e}")

        tk.Button(frame_botones, text="Cancelar", command=cancelar,
                  bg='#888888', fg='white', font=('Arial', 11, 'bold'),
                  padx=20, pady=6, cursor='hand2').pack(side='left')

        tk.Button(frame_botones, text="\u2709  ENVIAR CORREO", command=enviar,
                  bg='#D44638', fg='white', font=('Arial', 11, 'bold'),
                  padx=20, pady=6, cursor='hand2').pack(side='right')

    def generar(self):
        if not self.pagos:
            messagebox.showwarning("Sin datos", "No hay pagos registrados.")
            return
        acumulado = generar_excel(self.pagos)
        messagebox.showinfo("Excel generado",
                           f"Archivo generado: pagos_liliana.xlsx\n"
                           f"Total: ${acumulado:,.0f}\n"
                           f"Pagos: {len(self.pagos)}")


if __name__ == '__main__':
    root = tk.Tk()
    app = AppPagos(root)
    root.mainloop()
